from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file, jsonify
from db import crear_base_datos, insertar_empleado, cargar_empleado, existe_codigo
from qr import generar_qr
from imagen import generar_carnet, combinar_anverso_reverso
from procesador_fotos import procesar_foto_aprendiz
from datetime import date, timedelta, datetime
import os
import random
import traceback
import pandas as pd
import tempfile
import openpyxl
from werkzeug.utils import secure_filename
import sqlite3
import shutil
import json
import time
import re
import calendar

app = Flask(__name__)
app.secret_key = 'clave_secreta_segura'

# Configuraciones para Excel
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Crear carpetas necesarias
os.makedirs("static/fotos", exist_ok=True)
os.makedirs("static/qr", exist_ok=True)
os.makedirs("static/carnets", exist_ok=True)
os.makedirs("uploads", exist_ok=True)
os.makedirs("templates", exist_ok=True)
os.makedirs("static/fotos_backup", exist_ok=True)
os.makedirs("static/fotos_backup/por_fecha", exist_ok=True)
os.makedirs("static/fotos_backup/metadatos", exist_ok=True)

# Crear base de datos
crear_base_datos()

# Usuarios del sistema
usuarios = {
    "admin": {"clave": "admin123", "rol": "admin"},
    "aprendiz": {"clave": "aprendiz123", "rol": "aprendiz"},
    "sena": {"clave": "sena2024", "rol": "admin"},
    "usuario": {"clave": "123456", "rol": "admin"}
}

# TAREA DE VALIDACION - 17
# Vincula y recupera la cédula autorizada de la sesión del aprendiz.
def obtener_cedula_aprendiz_autenticado():
    """Devuelve la cédula autorizada para el aprendiz autenticado."""
    if session.get('rol') != 'aprendiz':
        return None

    cedula_sesion = session.get('aprendiz_cedula_auth')
    if cedula_sesion:
        return ''.join(filter(str.isdigit, str(cedula_sesion)))

    # Fallback: si el usuario de login es una cédula, se usa como identidad.
    usuario = str(session.get('usuario', '')).strip()
    usuario_limpio = ''.join(filter(str.isdigit, usuario))
    if 7 <= len(usuario_limpio) <= 12:
        session['aprendiz_cedula_auth'] = usuario_limpio
        return usuario_limpio

    return None


def obtener_archivo_carnet_por_cedula(cedula):
    """Retorna el nombre de archivo del carnet disponible para una cédula."""
    cedula_limpia = ''.join(filter(str.isdigit, str(cedula)))
    if not cedula_limpia:
        return None

    aprendiz = buscar_empleado_completo(cedula_limpia)
    if not aprendiz:
        return None

    # TAREA DE VALIDACION - 17
    # Prioriza archivo completo/combinado (anverso + reverso) antes del anverso solo.
    posibles_carnets = [
        os.path.join('static', 'carnets', f"{aprendiz['nombre'].replace(' ', '_')}_completo.png"),
        os.path.join('static', 'carnets', f'carnet_combinado_{cedula_limpia}.png'),
        os.path.join('static', 'carnets', f'carnet_{cedula_limpia}.png'),
    ]

    for carnet_path in posibles_carnets:
        if os.path.exists(carnet_path):
            return os.path.basename(carnet_path)
    return None

# =============================================
# FUNCIONES AUXILIARES PRINCIPALES
# =============================================

def actualizar_base_datos_sena():
    """Actualiza la base de datos con las columnas necesarias"""
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Verificar si la tabla existe
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='empleados'")
        if not cursor.fetchone():
            print("Creando tabla empleados...")
            # Crear tabla completa desde el inicio
            cursor.execute("""
                CREATE TABLE empleados (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nombre TEXT NOT NULL,
                    cedula TEXT UNIQUE NOT NULL,
                    tipo_documento TEXT DEFAULT 'CC',
                    cargo TEXT DEFAULT 'APRENDIZ',
                    codigo TEXT UNIQUE,
                    fecha_emision TEXT,
                    fecha_vencimiento TEXT,
                    tipo_sangre TEXT,
                    foto TEXT,
                    nis TEXT,
                    primer_apellido TEXT,
                    segundo_apellido TEXT,
                    nombre_programa TEXT,
                    codigo_ficha TEXT,
                    centro TEXT,
                    nivel_formacion TEXT DEFAULT 'Técnico',
                    red_tecnologica TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    
                    
                )
            """)
            print("Tabla empleados creada exitosamente")
        else:
            print("Tabla empleados existe, verificando columnas...")
            # Obtener columnas existentes
            cursor.execute("PRAGMA table_info(empleados)")
            columnas_existentes = [col[1] for col in cursor.fetchall()]
            
            # Columnas que deben existir
            columnas_necesarias = {
                'nis': 'TEXT',
                'primer_apellido': 'TEXT',
                'segundo_apellido': 'TEXT',
                'nombre_programa': 'TEXT',
                'codigo_ficha': 'TEXT',
                'centro': 'TEXT',
                'nivel_formacion': 'TEXT DEFAULT "Técnico"',
                'red_tecnologica': 'TEXT',
                'created_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
                'updated_at': 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP',
                'carnet_disponible': 'INTEGER DEFAULT 0'
            }
            
            # Agregar columnas faltantes
            for columna, tipo in columnas_necesarias.items():
                if columna not in columnas_existentes:
                    try:
                        cursor.execute(f'ALTER TABLE empleados ADD COLUMN {columna} {tipo}')
                        print(f"Columna agregada: {columna}")
                    except sqlite3.OperationalError as e:
                        if "duplicate column name" in str(e):
                            print(f"Columna ya existe: {columna}")
                        else:
                            print(f"Error agregando columna {columna}: {e}")
        
        # Crear índices para mejorar rendimiento
        indices = [
            "CREATE INDEX IF NOT EXISTS idx_cedula ON empleados(cedula)",
            "CREATE INDEX IF NOT EXISTS idx_codigo ON empleados(codigo)",
            "CREATE INDEX IF NOT EXISTS idx_nombre_programa ON empleados(nombre_programa)",
            "CREATE INDEX IF NOT EXISTS idx_codigo_ficha ON empleados(codigo_ficha)",
            "CREATE INDEX IF NOT EXISTS idx_fecha_emision ON empleados(fecha_emision)"
        ]
        
        for indice in indices:
            cursor.execute(indice)
        
        conn.commit()
        conn.close()
        print("Base de datos actualizada correctamente")
        return True
        
    except Exception as e:
        print(f"Error actualizando base de datos: {e}")
        return False

def buscar_empleado_completo(cedula):
    """Busca un empleado por cédula con todos los campos SENA"""
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Limpiar cédula de entrada
        cedula_limpia = ''.join(filter(str.isdigit, str(cedula)))
        
        print(f"Buscando empleado con cédula: {cedula_limpia}")
        
        cursor.execute("""
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE cedula = ? 
            ORDER BY created_at DESC, updated_at DESC
            LIMIT 1
        """, (cedula_limpia,))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            empleado = {
                'nombre': row[0] or '',
                'cedula': row[1] or '',
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4] or '',
                'fecha_emision': row[5] or '',
                'fecha_vencimiento': row[6] or '',
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8] or None,
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa Técnico',
                'codigo_ficha': row[13] or 'N/A',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Tecnologías de Producción Industrial'
            }
            
            print(f"Empleado encontrado: {empleado['nombre']} - Programa: {empleado['nombre_programa']}")
            return empleado
        else:
            print(f"No se encontró empleado con cédula: {cedula_limpia}")
            return None
            
    except Exception as e:
        print(f"Error buscando empleado: {e}")
        traceback.print_exc()
        return None

def obtener_todos_empleados():
    """Función para obtener todos los empleados de la base de datos"""
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados
            ORDER BY created_at DESC, nombre ASC
        """)
        
        empleados = []
        for row in cursor.fetchall():
            empleado = {
                'nombre': row[0] or '',
                'cedula': row[1] or '',
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4] or '',
                'fecha_emision': row[5] or '',
                'fecha_vencimiento': row[6] or '',
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8] or None,
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa Técnico',
                'codigo_ficha': row[13] or 'N/A',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Tecnologías de Producción Industrial'
            }
            empleados.append(empleado)
        
        conn.close()
        print(f"Obtenidos {len(empleados)} empleados de la base de datos")
        return empleados
        
    except Exception as e:
        print(f"Error obteniendo empleados: {e}")
        return []

def buscar_empleados_con_filtros(buscar='', filtro_foto='', filtro_programa='', filtro_nivel=''):
    """Busca empleados con múltiples filtros"""
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Construir query base
        query = """
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros
        if buscar:
            query += " AND (nombre LIKE ? OR cedula LIKE ? OR codigo LIKE ? OR nis LIKE ?)"
            buscar_param = f"%{buscar}%"
            params.extend([buscar_param, buscar_param, buscar_param, buscar_param])
        
        if filtro_foto == 'con_foto':
            query += " AND foto IS NOT NULL AND foto != ''"
        elif filtro_foto == 'sin_foto':
            query += " AND (foto IS NULL OR foto = '')"
        
        if filtro_programa:
            query += " AND nombre_programa LIKE ?"
            params.append(f"%{filtro_programa}%")
        
        if filtro_nivel:
            query += " AND nivel_formacion = ?"
            params.append(filtro_nivel)
        
        query += " ORDER BY nombre ASC"
        
        print(f"Ejecutando query: {query}")
        print(f"Con parámetros: {params}")
        
        cursor.execute(query, params)
        empleados = []
        
        for row in cursor.fetchall():
            empleado = {
                'nombre': row[0],
                'cedula': row[1],
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4],
                'fecha_emision': row[5],
                'fecha_vencimiento': row[6],
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8],
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa General',
                'codigo_ficha': row[13] or 'Sin Ficha',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Tecnologías de Producción Industrial'
            }
            empleados.append(empleado)
        
        conn.close()
        print(f"Encontrados {len(empleados)} empleados con los filtros aplicados")
        return empleados
        
    except Exception as e:
        print(f"Error buscando empleados con filtros: {e}")
        return []

def obtener_estadisticas_dashboard():
    """Obtiene estadísticas actualizadas para el dashboard"""
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Total de aprendices
        cursor.execute("SELECT COUNT(*) FROM empleados")
        total_aprendices = cursor.fetchone()[0]
        
        # Registrados hoy
        hoy = date.today().strftime("%Y-%m-%d")
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE fecha_emision = ?", (hoy,))
        registrados_hoy = cursor.fetchone()[0]
        
        # Esta semana
        fecha_semana = (date.today() - timedelta(days=7)).strftime("%Y-%m-%d")
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE fecha_emision >= ?", (fecha_semana,))
        esta_semana = cursor.fetchone()[0]
        
        # Con foto
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != ''")
        con_foto = cursor.fetchone()[0]
        
        # Por nivel de formación
        cursor.execute("SELECT nivel_formacion, COUNT(*) FROM empleados GROUP BY nivel_formacion")
        por_nivel = dict(cursor.fetchall())
        
        # Por programa
        cursor.execute("SELECT nombre_programa, COUNT(*) FROM empleados GROUP BY nombre_programa ORDER BY COUNT(*) DESC LIMIT 5")
        top_programas = cursor.fetchall()
        
        # Por ficha
        cursor.execute("SELECT codigo_ficha, COUNT(*) FROM empleados GROUP BY codigo_ficha ORDER BY COUNT(*) DESC LIMIT 5")
        top_fichas = cursor.fetchall()
        
        conn.close()
        
        return {
            'total_aprendices': total_aprendices,
            'registrados_hoy': registrados_hoy,
            'esta_semana': esta_semana,
            'con_foto': con_foto,
            'sin_foto': total_aprendices - con_foto,
            'por_nivel': por_nivel,
            'top_programas': top_programas,
            'top_fichas': top_fichas,
            'disponibilidad': 100 if total_aprendices > 0 else 0
        }
        
    except Exception as e:
        print(f"Error obteniendo estadísticas: {e}")
        return {
            'total_aprendices': 0,
            'registrados_hoy': 0,
            'esta_semana': 0,
            'con_foto': 0,
            'sin_foto': 0,
            'por_nivel': {},
            'top_programas': [],
            'top_fichas': [],
            'disponibilidad': 100
        }

def convertir_fecha_excel(fecha_serial):
    """Convierte fecha serial de Excel a formato YYYY-MM-DD"""
    try:
        if fecha_serial == "" or fecha_serial is None:
            return ""
        
        # Si ya es una cadena de fecha, devolverla
        if isinstance(fecha_serial, str):
            if "/" in fecha_serial or "-" in fecha_serial:
                return fecha_serial
        
        # Convertir número serial de Excel a fecha
        fecha_serial = float(fecha_serial)
        # Excel cuenta desde 1900-01-01, pero tiene un bug que cuenta 1900 como año bisiesto
        base_date = datetime(1899, 12, 30)  # Ajuste por el bug de Excel
        fecha_convertida = base_date + timedelta(days=fecha_serial)
        return fecha_convertida.strftime("%Y-%m-%d")
        
    except (ValueError, TypeError):
        return ""

def generar_nis_automatico():
    """Genera un NIS automático de 11 dígitos"""
    return str(random.randint(10000000000, 99999999999))

def determinar_nivel_formacion(programa):
    """Determina el nivel de formación basado en el programa"""
    programa_lower = programa.lower() if programa else ""
    
    # Palabras clave para tecnólogo
    tecnologicas = ["tecnología", "tecnológico", "tecnólogo", "gestión", "desarrollo", "análisis"]
    
    # Palabras clave para técnico
    tecnicas = ["técnico", "auxiliar", "operación", "mantenimiento"]
    
    for palabra in tecnologicas:
        if palabra in programa_lower:
            return "Tecnólogo"
    
    for palabra in tecnicas:
        if palabra in programa_lower:
            return "Técnico"
    
    # Por defecto, si el programa es largo (más de 50 caracteres), probablemente sea tecnólogo
    if len(programa) > 50:
        return "Tecnólogo"
    
    return "Técnico"

def procesar_foto_aprendiz_fallback(archivo_foto, cedula):
    """Función de procesamiento de fotos de respaldo si no existe la original"""
    try:
        # Validar tipo de archivo
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
        filename = archivo_foto.filename.lower()
        
        if not any(filename.endswith(ext) for ext in allowed_extensions):
            return False, None, "Formato de archivo no válido. Use PNG, JPG, JPEG o GIF"
        
        # Generar nombre único
        extension = filename.split('.')[-1]
        nombre_archivo = f"foto_{cedula}.{extension}"
        
        # Guardar archivo
        ruta_completa = os.path.join('static/fotos', nombre_archivo)
        archivo_foto.save(ruta_completa)
        
        return True, nombre_archivo, "Foto guardada correctamente"
        
    except Exception as e:
        return False, None, f"Error procesando foto: {str(e)}"

# =============================================
# FUNCIONES DE BACKUP DE FOTOS
# =============================================

def crear_carpetas_backup():
    """Crear carpetas de respaldo para fotos"""
    carpetas_backup = [
        "static/fotos_backup",
        "static/fotos_backup/por_fecha", 
        "static/fotos_backup/metadatos"
    ]
    
    for carpeta in carpetas_backup:
        if not os.path.exists(carpeta):
            os.makedirs(carpeta, exist_ok=True)
            print(f"Carpeta de backup creada: {carpeta}")

def crear_backup_foto(archivo_foto_path, cedula, usuario_tipo="admin", metadatos_extra=None):
    """
    Crea una copia de respaldo de la foto con metadatos
    
    Args:
        archivo_foto_path (str): Ruta del archivo foto original
        cedula (str): Cédula del aprendiz
        usuario_tipo (str): Tipo de usuario que subió ('admin' o 'aprendiz')
        metadatos_extra (dict): Datos adicionales a guardar
    """
    try:
        if not os.path.exists(archivo_foto_path):
            print(f"Archivo no encontrado para backup: {archivo_foto_path}")
            return False
        
        # Generar timestamp único
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fecha_carpeta = datetime.now().strftime("%Y-%m")
        
        # Crear carpeta por fecha si no existe
        backup_fecha_dir = f"static/fotos_backup/por_fecha/{fecha_carpeta}"
        os.makedirs(backup_fecha_dir, exist_ok=True)
        
        # Generar nombre único para el backup
        extension = os.path.splitext(archivo_foto_path)[1]
        nombre_backup = f"backup_{cedula}_{timestamp}{extension}"
        ruta_backup = os.path.join(backup_fecha_dir, nombre_backup)
        
        # Copiar archivo
        shutil.copy2(archivo_foto_path, ruta_backup)
        
        # Crear metadatos
        metadatos = {
            "cedula": cedula,
            "timestamp": timestamp,
            "fecha_backup": datetime.now().isoformat(),
            "archivo_original": archivo_foto_path,
            "archivo_backup": ruta_backup,
            "usuario_tipo": usuario_tipo,
            "tamano_bytes": os.path.getsize(archivo_foto_path)
        }
        
        if metadatos_extra:
            metadatos.update(metadatos_extra)
        
        # Guardar metadatos
        metadatos_file = f"static/fotos_backup/metadatos/backup_{cedula}_{timestamp}.json"
        with open(metadatos_file, 'w', encoding='utf-8') as f:
            json.dump(metadatos, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Backup creado: {nombre_backup}")
        return True
        
    except Exception as e:
        print(f"❌ Error creando backup: {e}")
        return False

def procesar_foto_aprendiz_con_backup(archivo_foto, cedula):
    """Versión modificada que incluye backup automático"""
    try:
        # Procesar foto normalmente
        exito, nombre_archivo_foto, mensaje = procesar_foto_aprendiz(archivo_foto, cedula)
        
        if exito:
            # Crear backup de la foto procesada
            ruta_foto_principal = os.path.join('static/fotos', nombre_archivo_foto)
            
            # Obtener datos del aprendiz para metadatos
            aprendiz = buscar_empleado_completo(cedula)
            metadatos_extra = {
                "nombre_aprendiz": aprendiz.get('nombre', 'N/A') if aprendiz else 'N/A',
                "programa": aprendiz.get('nombre_programa', 'N/A') if aprendiz else 'N/A',
                "codigo_ficha": aprendiz.get('codigo_ficha', 'N/A') if aprendiz else 'N/A'
            }
            
            # Crear backup
            backup_exitoso = crear_backup_foto(
                ruta_foto_principal, 
                cedula, 
                usuario_tipo="aprendiz",
                metadatos_extra=metadatos_extra
            )
            
            if backup_exitoso:
                mensaje += " (Copia de respaldo creada)"
            
        return exito, nombre_archivo_foto, mensaje
        
    except Exception as e:
        # Si falla el backup, usar la función original
        return procesar_foto_aprendiz_fallback(archivo_foto, cedula)

def procesar_foto_admin_con_backup(archivo_foto, cedula):
    """Función para cuando el admin sube fotos (también con backup)"""
    try:
        # Procesar foto normalmente
        exito, nombre_archivo_foto, mensaje = procesar_foto_aprendiz(archivo_foto, cedula)
        
        if exito:
            # Crear backup
            ruta_foto_principal = os.path.join('static/fotos', nombre_archivo_foto)
            
            aprendiz = buscar_empleado_completo(cedula)
            metadatos_extra = {
                "nombre_aprendiz": aprendiz.get('nombre', 'N/A') if aprendiz else 'N/A',
                "programa": aprendiz.get('nombre_programa', 'N/A') if aprendiz else 'N/A',
                "subida_por": "admin"
            }
            
            crear_backup_foto(
                ruta_foto_principal, 
                cedula, 
                usuario_tipo="admin",
                metadatos_extra=metadatos_extra
            )
            
        return exito, nombre_archivo_foto, mensaje
        
    except Exception as e:
        return procesar_foto_aprendiz_fallback(archivo_foto, cedula)

# =============================================
# NUEVA FUNCIÓN PARA VERIFICAR DUPLICADOS
# =============================================

def verificar_datos_duplicados(cedulas_excel, porcentaje_minimo=80):
    """
    Verifica si los datos del Excel ya están cargados en la base de datos
    Retorna True si más del porcentaje_minimo de las cédulas ya existen
    """
    try:
        if not cedulas_excel:
            return False, 0, 0
            
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Convertir cédulas a lista para consulta SQL
        cedulas_limpias = []
        for cedula in cedulas_excel:
            cedula_limpia = ''.join(filter(str.isdigit, str(cedula))) if cedula else ''
            if cedula_limpia and len(cedula_limpia) >= 7:
                cedulas_limpias.append(cedula_limpia)
        
        if not cedulas_limpias:
            conn.close()
            return False, 0, 0
        
        # Buscar cédulas existentes en la base de datos
        placeholders = ','.join(['?' for _ in cedulas_limpias])
        cursor.execute(f"SELECT cedula FROM empleados WHERE cedula IN ({placeholders})", cedulas_limpias)
        cedulas_existentes = [row[0] for row in cursor.fetchall()]
        
        conn.close()
        
        # Calcular porcentaje de coincidencias
        total_cedulas = len(cedulas_limpias)
        cedulas_encontradas = len(cedulas_existentes)
        porcentaje_coincidencia = (cedulas_encontradas / total_cedulas) * 100 if total_cedulas > 0 else 0
        
        print(f"📊 Verificación duplicados:")
        print(f"   - Total cédulas en Excel: {total_cedulas}")
        print(f"   - Cédulas ya existentes: {cedulas_encontradas}")
        print(f"   - Porcentaje coincidencia: {porcentaje_coincidencia:.1f}%")
        
        # Si más del porcentaje_minimo ya existe, consideramos que está duplicado
        es_duplicado = porcentaje_coincidencia >= porcentaje_minimo
        
        return es_duplicado, cedulas_encontradas, total_cedulas
        
    except Exception as e:
        print(f"Error verificando duplicados: {e}")
        return False, 0, 0

# =============================================
# FUNCIÓN MEJORADA PARA CARGAR EXCEL SENA (CON VERIFICACIÓN DE DUPLICADOS)
# =============================================

def cargar_excel_sena_mejorado(file):
    """Función especializada para cargar archivos Excel del SENA con manejo mejorado"""
    try:
        print("=== INICIANDO CARGA DE EXCEL SENA ===")
        
        # Guardar archivo temporalmente
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            file.save(temp_file.name)
            temp_file_path = temp_file.name
        
        print(f"Archivo guardado temporalmente en: {temp_file_path}")
        
        # Leer Excel con openpyxl directamente para mejor control
        workbook = openpyxl.load_workbook(temp_file_path)
        sheet = workbook.active
        
        print(f"Hoja activa: {sheet.title}")
        
        # Obtener todas las filas
        rows = list(sheet.iter_rows(values_only=True))
        print(f"Total de filas encontradas: {len(rows)}")
        
        if len(rows) < 2:
            os.unlink(temp_file_path)
            return {'success': False, 'message': 'El archivo no contiene datos válidos'}
        
        # Identificar headers (primera fila no vacía)
        headers_row = rows[0]
        print(f"Headers detectados: {headers_row}")
        
        # Crear mapa de columnas (ignorando las vacías)
        column_map = {}
        for idx, header in enumerate(headers_row):
            if header and header.strip():
                column_map[header.strip()] = idx
        
        print(f"Mapa de columnas creado: {column_map}")
        print(f"🔍 Buscando columna de fecha en: {list(column_map.keys())}")
        
        # Verificar columnas requeridas
        required_columns = ['Primer Apellido', 'Nombre', 'Tipo de documento', 'Número de documento']
        missing_columns = [col for col in required_columns if col not in column_map]
        
        if missing_columns:
            os.unlink(temp_file_path)
            return {
                'success': False, 
                'message': f'Faltan columnas requeridas: {", ".join(missing_columns)}'
            }
        
        # 🆕 NUEVA VERIFICACIÓN DE DUPLICADOS
        print("🔍 Verificando si los datos ya están cargados...")
        
        # Extraer todas las cédulas del Excel para verificación
        cedulas_excel = []
        for row_idx, row in enumerate(rows[1:], start=2):
            if not any(row):
                continue
            numero_documento = str(row[column_map.get('Número de documento', '')]).strip() if row[column_map.get('Número de documento', '')] else ''
            if numero_documento and numero_documento != 'None':
                cedulas_excel.append(numero_documento)
        
        # Verificar duplicados
        es_duplicado, coincidencias, total = verificar_datos_duplicados(cedulas_excel, porcentaje_minimo=80)
        
        if es_duplicado:
            os.unlink(temp_file_path)
            return {
                'success': False,
                'message': f'⚠️ Base de datos duplicada detectada. {coincidencias} de {total} aprendices ({((coincidencias/total)*100):.1f}%) ya están registrados en el sistema. Si necesitas actualizar datos específicos, elimina los registros duplicados primero o usa una plantilla con solo los datos nuevos.',
                'duplicado': True,
                'coincidencias': coincidencias,
                'total': total
            }
        
        # Si llegamos aquí, continuar con el procesamiento normal...
        print("✅ Verificación pasada, continuando con la carga...")
        
        # Procesar datos
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        # Procesar cada fila de datos (saltar header)
        for row_idx, row in enumerate(rows[1:], start=2):
            try:
                # Saltar filas completamente vacías
                if not any(row):
                    continue
                
                # Extraer datos usando el mapa de columnas
                numero_documento = str(row[column_map.get('Número de documento', '')]).strip() if row[column_map.get('Número de documento', '')] else ''
                
                # Validar que tenga número de documento
                if not numero_documento or numero_documento == 'None':
                    continue
                
                # Limpiar número de documento
                numero_documento = ''.join(filter(str.isdigit, numero_documento))
                if len(numero_documento) < 7:
                    errors.append(f"Fila {row_idx}: Número de documento inválido")
                    error_count += 1
                    continue
                
                # Extraer otros campos
                primer_apellido = str(row[column_map.get('Primer Apellido', '')]).strip().upper() if row[column_map.get('Primer Apellido', '')] else ''
                segundo_apellido = str(row[column_map.get('Segundo Apellido', '')]).strip().upper() if row[column_map.get('Segundo Apellido', '')] else ''
                nombre = str(row[column_map.get('Nombre', '')]).strip().upper() if row[column_map.get('Nombre', '')] else ''
                tipo_documento = str(row[column_map.get('Tipo de documento', '')]).strip() if row[column_map.get('Tipo de documento', '')] else 'CC'
                tipo_sangre = str(row[column_map.get('Tipo de Sangre', '')]).strip().upper() if row[column_map.get('Tipo de Sangre', '')] else 'O+'
                nombre_programa = str(row[column_map.get('Nombre del Programa', '')]).strip() if row[column_map.get('Nombre del Programa', '')] else ''
                codigo_ficha = str(row[column_map.get('Código de Ficha', '')]).strip() if row[column_map.get('Código de Ficha', '')] else ''
                centro = str(row[column_map.get('Centro', '')]).strip() if row[column_map.get('Centro', '')] else 'Centro de Biotecnología Industrial'
                red_tecnologica = str(row[column_map.get('Red Tecnologica', '')]).strip() if row[column_map.get('Red Tecnologica', '')] else ''
                
                # Procesar fecha
                fecha_finalizacion = None
                posibles_nombres_fecha = [
                    'Fecha Finalización del Programa',
                    'Fecha Finalizacion del Programa',
                    'FECHA FINALIZACION',
                    'Fecha de Finalización',
                    'Fecha Finalizacion',
                    'Fecha Final',
                    'Fecha Fin'
                ]

                for nombre_col in posibles_nombres_fecha:
                    if nombre_col in column_map:
                        fecha_finalizacion_raw = row[column_map.get(nombre_col, '')]
                        if fecha_finalizacion_raw and fecha_finalizacion_raw != 'None':
                            fecha_finalizacion = convertir_fecha_excel(fecha_finalizacion_raw)
                            print(f"✅ Fecha encontrada en columna '{nombre_col}': {fecha_finalizacion}")
                            break

                # Si no se encontró fecha, usar fecha por defecto (1 año desde hoy)
                if not fecha_finalizacion:
                    from datetime import datetime, timedelta
                    fecha_default = (datetime.now() + timedelta(days=365)).strftime("%Y-%m-%d")
                    fecha_finalizacion = fecha_default
                    print(f"⚠️ No se encontró fecha de finalización, usando: {fecha_finalizacion}")
                
                # Generar o procesar NIS
                nis = str(row[column_map.get('NIS', '')]).strip() if row[column_map.get('NIS', '')] else ''
                if not nis or nis == 'None' or nis == '':
                    nis = generar_nis_automatico()
                
                # Determinar nivel de formación
                nivel_formacion = determinar_nivel_formacion(nombre_programa)
                
                # Validar datos mínimos
                if not all([primer_apellido, nombre, numero_documento]):
                    errors.append(f"Fila {row_idx}: Faltan datos obligatorios (Primer Apellido, Nombre, Número de documento)")
                    error_count += 1
                    continue
                
                # Construir nombre completo
                nombre_completo = f"{nombre} {primer_apellido}"
                if segundo_apellido:
                    nombre_completo += f" {segundo_apellido}"
                
                # Verificar si ya existe
                cursor.execute("SELECT id FROM empleados WHERE cedula = ?", (numero_documento,))
                existe = cursor.fetchone()
                
                # Generar código único
                iniciales = ''.join([parte[0] for parte in nombre_completo.split() if parte])[:4]
                codigo_generado = None
                for _ in range(10):
                    codigo_temp = f"{iniciales}{random.randint(1000, 9999)}"
                    cursor.execute("SELECT codigo FROM empleados WHERE codigo = ?", (codigo_temp,))
                    if not cursor.fetchone():
                        codigo_generado = codigo_temp
                        break
                
                if not codigo_generado:
                    errors.append(f"Fila {row_idx}: No se pudo generar código único")
                    error_count += 1
                    continue
                
                # Preparar datos
                hoy = date.today().strftime("%Y-%m-%d")
                
                if existe:
                    # Actualizar
                    cursor.execute("""
                        UPDATE empleados SET 
                            nombre = ?, tipo_documento = ?, cargo = ?, codigo = ?,
                            fecha_emision = ?, fecha_vencimiento = ?, tipo_sangre = ?,
                            nis = ?, primer_apellido = ?, segundo_apellido = ?,
                            nombre_programa = ?, codigo_ficha = ?, centro = ?, 
                            nivel_formacion = ?, red_tecnologica = ?, updated_at = CURRENT_TIMESTAMP
                        WHERE cedula = ?
                    """, (
                        nombre_completo, tipo_documento, 'APRENDIZ', codigo_generado,
                        hoy, fecha_finalizacion, tipo_sangre,
                        nis, primer_apellido, segundo_apellido,
                        nombre_programa, codigo_ficha, centro, 
                        nivel_formacion, red_tecnologica,
                        numero_documento
                    ))
                    updated_count += 1
                    print(f"Actualizado: {nombre_completo} - Cedula: {numero_documento}")
                else:
                    # Crear nuevo
                    cursor.execute("""
                        INSERT INTO empleados (
                            nombre, cedula, tipo_documento, cargo, codigo,
                            fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                            nis, primer_apellido, segundo_apellido,
                            nombre_programa, codigo_ficha, centro, 
                            nivel_formacion, red_tecnologica
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        nombre_completo, numero_documento, tipo_documento, 
                        'APRENDIZ', codigo_generado, hoy, 
                        fecha_finalizacion, tipo_sangre, None,
                        nis, primer_apellido, segundo_apellido,
                        nombre_programa, codigo_ficha, centro, 
                        nivel_formacion, red_tecnologica
                    ))
                    created_count += 1
                    print(f"Creado: {nombre_completo} - Cedula: {numero_documento}")
                
                # Confirmar cada inserción
                conn.commit()
                
            except Exception as e:
                error_count += 1
                error_msg = f"Fila {row_idx}: Error - {str(e)}"
                errors.append(error_msg)
                print(error_msg)
                continue
        
        conn.close()
        os.unlink(temp_file_path)
        
        print(f"=== CARGA COMPLETADA ===")
        print(f"Creados: {created_count}")
        print(f"Actualizados: {updated_count}")
        print(f"Errores: {error_count}")
        
        return {
            'success': True,
            'created': created_count,
            'updated': updated_count,
            'errors': error_count,
            'error_details': errors[:10],  # Máximo 10 errores para mostrar
            'message': f'✅ Carga exitosa del SENA: {created_count} aprendices creados, {updated_count} actualizados.'
        }
        
    except Exception as e:
        if 'temp_file_path' in locals():
            os.unlink(temp_file_path)
        print(f"Error general cargando Excel: {e}")
        return {
            'success': False,
            'message': f'Error al procesar archivo: {str(e)}'
        }

# =============================================
# RUTAS PRINCIPALES DEL SISTEMA
# =============================================

@app.route('/')
def index():
    # Si ya está logueado, redirigir según el rol
    if 'usuario' in session:
        if session.get('rol') == 'admin':
            return redirect(url_for('dashboard_admin'))
        elif session.get('rol') == 'aprendiz':
            return redirect(url_for('dashboard_aprendiz'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Mejorado para soportar ambos nombres de campo
        usuario = request.form.get('usuario', '').strip() or request.form.get('username', '').strip()
        clave = request.form.get('password', '').strip() or request.form.get('clave', '').strip()

        print(f"Intento de login - Usuario: {usuario}, Clave: {clave}")

        # Validación mejorada con múltiples credenciales
        if usuario in usuarios and usuarios[usuario]["clave"] == clave:
            session['usuario'] = usuario
            session['rol'] = usuarios[usuario]["rol"]
            flash(f'Bienvenido {usuario}! Has iniciado sesión correctamente.', 'success')
            
            # Redirigir según el rol
            if session['rol'] == 'admin':
                return redirect(url_for('dashboard_admin'))
            elif session['rol'] == 'aprendiz':
                return redirect(url_for('dashboard_aprendiz'))
        else:
            flash("Usuario o contraseña incorrectos. Intenta de nuevo.", 'error')
            return render_template('login.html', error='Credenciales incorrectas')

    return render_template('login.html')


@app.route('/api/metricas_dashboard')
def api_metricas_dashboard():
    """Métricas en tiempo real para el dashboard dinámico"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403

    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()

        hoy = date.today().strftime("%Y-%m-%d")
        semana = (date.today() - timedelta(days=7)).strftime("%Y-%m-%d")

        # ── Aprendices ──
        cursor.execute("SELECT COUNT(*) FROM empleados")
        total = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE DATE(created_at) = ?", (hoy,))
        hoy_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE DATE(created_at) >= ?", (semana,))
        semana_count = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != ''")
        con_foto = cursor.fetchone()[0]

        sin_foto = total - con_foto

        conn.close()

        # ── Carnets generados (leyendo disco) ──
        carpeta = os.path.join('static', 'carnets')
        cedulas_carnet = set()

        if os.path.exists(carpeta):
            archivos_completo = []
            for archivo in os.listdir(carpeta):
                if not archivo.endswith('.png'):
                    continue
                if archivo.startswith('carnet_') and '_completo' not in archivo:
                    ced = archivo.replace('carnet_', '').replace('.png', '')
                    if ced.isdigit():
                        cedulas_carnet.add(ced)
                elif archivo.startswith('carnet_combinado_'):
                    ced = archivo.replace('carnet_combinado_', '').replace('.png', '')
                    if ced.isdigit():
                        cedulas_carnet.add(ced)
                if archivo.endswith('_completo.png'):
                    archivos_completo.append(archivo)

            if archivos_completo:
                conn2 = sqlite3.connect('carnet.db')
                cur2 = conn2.cursor()
                cur2.execute("SELECT cedula, nombre FROM empleados")
                for ced_bd, nom_bd in cur2.fetchall():
                    if nom_bd and (nom_bd.replace(' ', '_') + '_completo.png') in archivos_completo:
                        cedulas_carnet.add(ced_bd)
                conn2.close()

        # Cruzar con cédulas actuales en BD para no contar huérfanos
        conn3 = sqlite3.connect('carnet.db')
        cur3 = conn3.cursor()
        cur3.execute("SELECT cedula, foto FROM empleados")
        rows = cur3.fetchall()
        conn3.close()

        cedulas_bd    = {r[0] for r in rows}
        cedulas_foto  = {r[0] for r in rows if r[1]}

        carnets_validos   = cedulas_carnet & cedulas_bd          # con carnet en disco Y en BD
        carnets_con_foto  = carnets_validos & cedulas_foto       # generados Y con foto actual
        carnets_sin_foto  = carnets_validos - cedulas_foto       # generados pero sin foto
        total_carnets     = len(carnets_validos)
        pendientes        = len(cedulas_foto - carnets_validos)  # tienen foto pero falta generar

        # Carnets de hoy / semana — aproximación por fecha de archivo
        carnets_hoy    = 0
        carnets_semana = 0
        if os.path.exists(carpeta):
            import time as _time
            ts_hoy   = _time.time() - 86400
            ts_semana= _time.time() - 86400 * 7
            for archivo in os.listdir(carpeta):
                ruta = os.path.join(carpeta, archivo)
                if archivo.endswith('.png') and os.path.exists(ruta):
                    mtime = os.path.getmtime(ruta)
                    if mtime >= ts_hoy:
                        carnets_hoy += 1
                    if mtime >= ts_semana:
                        carnets_semana += 1

        return jsonify({
            'success': True,
            # Aprendices
            'total_aprendices': total,
            'registrados_hoy': hoy_count,
            'registrados_semana': semana_count,
            'con_foto': con_foto,
            'sin_foto': sin_foto,
            # Carnets
            'carnets_total': total_carnets,
            'carnets_hoy': carnets_hoy,
            'carnets_semana': carnets_semana,
            'carnets_con_foto': len(carnets_con_foto),
            'carnets_sin_foto': len(carnets_sin_foto),
            'carnets_pendientes': pendientes,
        })

    except Exception as e:
        print(f"Error en api_metricas_dashboard: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500




@app.route('/logout')
def logout():
    usuario = session.get('usuario', 'Usuario')
    session.clear()
    flash(f'Has cerrado sesión exitosamente. Hasta pronto {usuario}!', 'info')
    return redirect(url_for('login'))

@app.route('/logout', methods=['POST'])
def logout_post():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@app.route('/dashboard_admin')
def dashboard_admin():
    if 'usuario' not in session or session['rol'] != 'admin':
        flash('Debes iniciar sesión como administrador para acceder.', 'error')
        return redirect(url_for('login'))

    stats = obtener_estadisticas_dashboard()

    # Consultar carnets marcados como disponibles
    conn = sqlite3.connect('carnet.db')
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1")
    carnets_disponibles = cursor.fetchone()[0]
    cursor.execute("""
        SELECT nombre, cedula, codigo
        FROM empleados
        WHERE carnet_disponible = 1
    """)
    carnets = cursor.fetchall()
    conn.close()

    stats['carnets_disponibles'] = carnets_disponibles

    return render_template(
        "dashboard_admin.html",
        usuario=session['usuario'],
        stats=stats,
        carnets=carnets   # ESTA ES LA NUEVA VARIABLE
    )

# TAREA DE VALIDACION - 17
# Se envía al frontend la cédula autorizada para bloquear consulta en UI.
@app.route('/dashboard_aprendiz')
def dashboard_aprendiz():
    if 'usuario' not in session or session['rol'] != 'aprendiz':
        flash('Debes iniciar sesión como aprendiz para acceder.', 'error')
        return redirect(url_for('login'))
    return render_template(
        "dashboard_aprendiz.html",
        usuario=session['usuario'],
        cedula_autenticada=obtener_cedula_aprendiz_autenticado()
    )

# =====================================================
# 🆕 NUEVOS ENDPOINTS DE API PARA LA BÚSQUEDA
# =====================================================

@app.route('/api/lista_aprendices_filtrada', methods=['GET'])
def api_lista_aprendices_filtrada():
    """
    API para obtener lista filtrada de aprendices
    
    Parámetros:
    - todos=true: mostrar todos
    - ficha=codigo: buscar por ficha
    - cedula=numero: buscar por cédula exacta
    - nombre=texto: buscar por nombre (parcial)
    - foto=con_foto|sin_foto: filtrar por foto
    """
    try:
        print("\n[API] === BÚSQUEDA INICIADA ===")
        
        # Obtener parámetros
        todos = request.args.get('todos', '').lower() == 'true'
        ficha = request.args.get('ficha', '').strip()
        cedula = request.args.get('cedula', '').strip()
        nombre = request.args.get('nombre', '').strip()
        foto_estado = request.args.get('foto', '').strip()
        
        print(f"[API] Parámetros recibidos:")
        print(f"  - todos: {todos}")
        print(f"  - ficha: {ficha}")
        print(f"  - cedula: {cedula}")
        print(f"  - nombre: {nombre}")
        print(f"  - foto_estado: {foto_estado}")
        
        # Conectar a BD
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Contar total en BD
        cursor.execute("SELECT COUNT(*) FROM empleados")
        total_en_bd = cursor.fetchone()[0]
        print(f"[API] Total en BD: {total_en_bd}")
        
        # Construir consulta
        query = """
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros
        if not todos:
            if ficha:
                print(f"[API] Filtrando por ficha: {ficha}")
                query += " AND codigo_ficha LIKE ?"
                params.append(f"%{ficha}%")
            
            if cedula:
                print(f"[API] Filtrando por cedula: {cedula}")
                cedula_limpia = ''.join(filter(str.isdigit, cedula))
                query += " AND cedula = ?"
                params.append(cedula_limpia)
            
            if nombre:
                print(f"[API] Filtrando por nombre: {nombre}")
                query += " AND nombre LIKE ?"
                params.append(f"%{nombre}%")
        else:
            print(f"[API] Mostrando TODOS los aprendices")
        
        # Filtro por foto
        if foto_estado == 'con_foto':
            print(f"[API] Filtrando: CON FOTO")
            query += " AND foto IS NOT NULL AND foto != ''"
        elif foto_estado == 'sin_foto':
            print(f"[API] Filtrando: SIN FOTO")
            query += " AND (foto IS NULL OR foto = '')"
        
        query += " ORDER BY nombre ASC"
        
        print(f"[API] Query: {query}")
        print(f"[API] Parámetros: {params}")
        
        # Ejecutar
        cursor.execute(query, params)
        aprendices = []
        
        for row in cursor.fetchall():
            aprendiz = {
                'nombre': row[0],
                'cedula': row[1],
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4],
                'fecha_emision': row[5],
                'fecha_vencimiento': row[6],
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8],
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa General',
                'codigo_ficha': row[13] or 'Sin Ficha',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Red Tecnológica'
            }
            aprendices.append(aprendiz)
        
        conn.close()
        
        print(f"[API] RESULTADO: {len(aprendices)} aprendices encontrados")
        print(f"[API] === BÚSQUEDA FINALIZADA ===\n")
        
        return jsonify({
            'success': True,
            'total': len(aprendices),
            'aprendices': aprendices
        })
    
    except Exception as e:
        print(f"[API ERROR]: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500

# TAREA DE VALIDACION - 17
# API con control de propiedad: primera búsqueda fija cédula y luego la restringe.
@app.route('/api/buscar_aprendiz/<cedula>')
def api_buscar_aprendiz(cedula):
    """Buscar un aprendiz específico por cédula"""
    try:

        cedula_limpia = ''.join(filter(str.isdigit, cedula))

        if 'usuario' not in session:
            return jsonify({'success': False, 'message': 'Debes iniciar sesión para consultar datos'}), 401

        if session.get('rol') == 'aprendiz':
            cedula_autorizada = obtener_cedula_aprendiz_autenticado()
            if cedula_autorizada and cedula_limpia != cedula_autorizada:
                return jsonify({
                    'success': False,
                    'message': f'Solo puedes consultar la cédula vinculada a esta sesión: {cedula_autorizada}.'
                }), 403

        empleado = buscar_empleado_completo(cedula_limpia)
        
        if empleado:
            # ✅ VERIFICAR FOTO
            if empleado['foto']:
                ruta_foto = os.path.join('static/fotos', empleado['foto'])
                empleado['foto_existe'] = os.path.exists(ruta_foto)
                empleado['foto_url'] = f"/static/fotos/{empleado['foto']}" if empleado['foto_existe'] else None
            else:
                empleado['foto_existe'] = False
                empleado['foto_url'] = None

            # ✅ VERIFICAR CARNET (CORRECTO)
            carpeta_carnets = os.path.join(app.root_path, 'static', 'carnets')

            try:
                archivos = os.listdir(carpeta_carnets)
                empleado['carnet_generado'] = any(cedula_limpia in f for f in archivos)
            except Exception as e:
                print("Error leyendo carnets:", e)
                empleado['carnet_generado'] = False

            # Si es aprendiz y aún no tiene cédula vinculada, se fija en la primera consulta exitosa
            if session.get('rol') == 'aprendiz' and not obtener_cedula_aprendiz_autenticado():
                session['aprendiz_cedula_auth'] = cedula_limpia

            return jsonify({'success': True, 'data': empleado})
        else:
            return jsonify({'success': False, 'message': 'Aprendiz no encontrado'})
            
    except Exception as e:
        print(f"Error API buscar aprendiz: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

# =============================================
# RESTO DE RUTAS (copiar todo lo demás de tu archivo app.py original)
# =============================================

@app.route('/agregar', methods=['GET', 'POST'])
def agregar():

    print(f"RUTA AGREGAR ACCEDIDA - MÉTODO: {request.method}")
    
    if 'usuario' not in session or session['rol'] != 'admin':
        return redirect(url_for('login'))

    hoy = date.today()
    vencimiento = hoy + timedelta(days=365)

    if request.method == 'POST':
        try:
            print("PROCESANDO FORMULARIO...")
            print("DATOS RECIBIDOS:", dict(request.form))
            print("ARCHIVOS RECIBIDOS:", dict(request.files))
            
            # OBTENER CAMPOS BÁSICOS OBLIGATORIOS
            nis = request.form.get('nis', '').strip()
            primer_apellido = request.form.get('primer_apellido', '').strip().upper()
            segundo_apellido = request.form.get('segundo_apellido', '').strip().upper()
            nombres = request.form.get('nombres', '').strip().upper()
            tipo_documento = request.form.get('tipo_documento', '').strip()
            cedula = request.form.get('cedula', '').strip()
            tipo_sangre = request.form.get('tipo_sangre', '').strip().upper()
            fecha_vencimiento = request.form.get('fecha_vencimiento', '').strip()
            nombre_programa = request.form.get('nombre_programa', '').strip()
            codigo_ficha = request.form.get('codigo_ficha', '').strip()
            
            # NUEVO CAMPO - Nivel de Formación
            nivel_formacion = request.form.get('nivel_formacion', '').strip()
            
            print(f"CAMPOS EXTRAÍDOS: NIS={nis}, Nombres={nombres}, Nivel={nivel_formacion}")
            
            # VALIDACIONES BÁSICAS (ahora incluye nivel_formacion)
            if not all([nis, primer_apellido, nombres, tipo_documento, cedula, tipo_sangre, 
                       fecha_vencimiento, nombre_programa, codigo_ficha, nivel_formacion]):
                flash("Todos los campos obligatorios deben estar completos.", 'error')
                print("VALIDACIÓN FALLIDA - Campos faltantes")
                return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))
            
            # CONSTRUIR NOMBRE COMPLETO
            apellidos = f"{primer_apellido} {segundo_apellido}".strip()
            nombre_completo = f"{nombres} {apellidos}".strip()
            centro = "Centro de Biotecnología Industrial"
            cargo = 'APRENDIZ'
            
            print(f"NOMBRE COMPLETO: {nombre_completo}")
            
            # GENERAR CÓDIGO ÚNICO
            iniciales = ''.join([parte[0] for parte in nombre_completo.split() if parte])
            codigo = None
            for _ in range(10):
                codigo_temp = f"{iniciales}{random.randint(1000, 9999)}"
                if not existe_codigo(codigo_temp):
                    codigo = codigo_temp
                    break
            
            if not codigo:
                flash("No se pudo generar un código único.", 'error')
                print("ERROR GENERANDO CÓDIGO")
                return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))
            
            print(f"CÓDIGO GENERADO: {codigo}")
            
            # MANEJAR FOTO OBLIGATORIA CON PROCESAMIENTO AUTOMÁTICO Y BACKUP
            archivo_foto = request.files.get('foto')
            nombre_archivo_foto = None
            
            if archivo_foto and archivo_foto.filename != '':
                # Usar función con backup para admin
                exito, nombre_archivo_foto, mensaje = procesar_foto_admin_con_backup(archivo_foto, cedula)
                
                if not exito:
                    flash(f"Error procesando la foto: {mensaje}", 'error')
                    print("ERROR PROCESANDO FOTO")
                    return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))
                
                print(f"FOTO PROCESADA: {nombre_archivo_foto}")
                flash("Foto procesada automáticamente: 3x4, fondo blanco, optimizada para carnet", 'info')
            else:
                flash("Debe subir una foto.", 'error')
                print("FOTO FALTANTE")
                return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))
            
            # PREPARAR DATOS PARA INSERTAR (CON NIVEL_FORMACION)
            datos = {
                'nombre': nombre_completo,
                'cedula': cedula,
                'tipo_documento': tipo_documento,
                'cargo': cargo,
                'codigo': codigo,
                'fecha_emision': hoy.strftime("%Y-%m-%d"),
                'fecha_vencimiento': fecha_vencimiento,
                'tipo_sangre': tipo_sangre,
                'foto': nombre_archivo_foto,
                'nis': nis,
                'primer_apellido': primer_apellido,
                'segundo_apellido': segundo_apellido,
                'nombre_programa': nombre_programa,
                'codigo_ficha': codigo_ficha,
                'centro': centro,
                'nivel_formacion': nivel_formacion
            }
            
            print("DATOS PREPARADOS PARA INSERTAR:")
            for key, value in datos.items():
                print(f"   {key}: {value}")
            
            # INSERTAR EN BASE DE DATOS
            print("INSERTANDO EN BASE DE DATOS...")
            insertar_empleado(datos)
            print("EMPLEADO INSERTADO CORRECTAMENTE")
            
            # MENSAJE DE ÉXITO Y REDIRECCIÓN
            flash(f"Empleado {nombre_completo} registrado correctamente! Nivel: {nivel_formacion}", 'success')
            return redirect(url_for('agregar'))
            
        except Exception as e:
            print(f"ERROR COMPLETO: {str(e)}")
            print(f"TRACEBACK: {traceback.format_exc()}")
            flash(f"Error al guardar: {str(e)}", 'error')
            return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))
    
    # GET REQUEST - MOSTRAR FORMULARIO
    print("MOSTRANDO FORMULARIO GET")
    return render_template('agregar.html', fecha_hoy=hoy.strftime("%Y-%m-%d"), fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"))

@app.route('/agregar_empleado', methods=['GET', 'POST'])
def agregar_empleado():
    """Nueva ruta para agregar empleados con todos los campos del SENA"""
    if 'usuario' not in session or session['rol'] != 'admin':
        flash('Debes iniciar sesión como administrador para acceder.', 'error')
        return redirect(url_for('login'))

    if request.method == 'POST':
        try:
            # Obtener todos los campos del formulario SENA
            nis = request.form.get('nis', '').strip()
            primer_apellido = request.form.get('primer_apellido', '').strip().upper()
            segundo_apellido = request.form.get('segundo_apellido', '').strip().upper()
            nombre = request.form.get('nombre', '').strip().upper()
            tipo_documento = request.form.get('tipo_documento', '').strip()
            numero_documento = request.form.get('numero_documento', '').strip()
            tipo_sangre = request.form.get('tipo_sangre', '').strip().upper()
            nombre_programa = request.form.get('nombre_programa', '').strip()
            codigo_ficha = request.form.get('codigo_ficha', '').strip()
            
            # NUEVO CAMPO - Nivel de Formación
            nivel_formacion = request.form.get('nivel_formacion', '').strip()
            
            # CENTRO FIJO
            centro = "Centro de Biotecnología Industrial"
            fecha_finalizacion = request.form.get('fecha_finalizacion', '').strip()
            
            # Validaciones básicas (ahora incluye nivel_formacion)
            if not all([nis, primer_apellido, nombre, tipo_documento, numero_documento, 
                       tipo_sangre, nombre_programa, codigo_ficha, fecha_finalizacion, nivel_formacion]):
                flash('Todos los campos obligatorios deben estar completos.', 'error')
                return render_template('agregar_empleado.html')
            
            # Construir nombre completo
            nombre_completo = f"{nombre} {primer_apellido}"
            if segundo_apellido:
                nombre_completo += f" {segundo_apellido}"
            
            # Generar código único si no se proporcionó
            codigo_generado = request.form.get('codigo', '').strip()
            if not codigo_generado:
                iniciales = ''.join([parte[0] for parte in nombre_completo.split() if parte])
                for _ in range(10):
                    codigo_generado = f"{iniciales}{random.randint(1000, 9999)}"
                    if not existe_codigo(codigo_generado):
                        break
                else:
                    flash("No se pudo generar un código único. Intente nuevamente.", 'error')
                    return render_template('agregar_empleado.html')
            
            # Preparar datos para la base de datos (compatible con estructura existente)
            hoy = date.today()
            datos = {
                'nombre': nombre_completo,
                'cedula': numero_documento,
                'tipo_documento': tipo_documento,
                'cargo': 'APRENDIZ',  # Por defecto para SENA
                'codigo': codigo_generado,
                'fecha_emision': hoy.strftime("%Y-%m-%d"),
                'fecha_vencimiento': fecha_finalizacion,
                'tipo_sangre': tipo_sangre,
                'foto': None,
                # Campos adicionales SENA
                'nis': nis,
                'primer_apellido': primer_apellido,
                'segundo_apellido': segundo_apellido,
                'nombre_programa': nombre_programa,
                'codigo_ficha': codigo_ficha,
                'centro': centro,
                'nivel_formacion': nivel_formacion
            }
            
            # Manejar foto con backup
            archivo_foto = request.files.get('foto')
            if archivo_foto and archivo_foto.filename != '':
                # Usar función con backup para admin
                exito, nombre_archivo_foto, mensaje = procesar_foto_admin_con_backup(archivo_foto, numero_documento)
                
                if exito:
                    datos['foto'] = nombre_archivo_foto
                else:
                    flash(f"Error procesando foto: {mensaje}", 'error')
                    return render_template('agregar_empleado.html')
            else:
                flash("Debe subir una foto.", 'error')
                return render_template('agregar_empleado.html')
            
            # Insertar en base de datos (usa tu función existente)
            insertar_empleado(datos)
            flash(f"Empleado {nombre_completo} registrado correctamente en el sistema SENA. Nivel: {nivel_formacion}", 'success')
            return redirect(url_for('dashboard_admin'))
            
        except ValueError as ve:
            flash(str(ve), 'error')
            return render_template('agregar_empleado.html')
        except Exception as e:
            print(f"Error en agregar_empleado: {e}")
            flash(f"Error inesperado: {e}", 'error')
            return render_template('agregar_empleado.html')
    
    # GET request - mostrar formulario
    return render_template('agregar_empleado.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro_aprendiz():
    if 'usuario' not in session or session['rol'] != 'aprendiz':
        return redirect(url_for('login'))

    hoy = date.today()
    vencimiento = hoy + timedelta(days=365)

    if request.method == 'POST':
        nombres = request.form['nombres'].strip().upper()
        apellidos = request.form['apellidos'].strip().upper()
        nombre = f"{nombres} {apellidos}"
        tipo_documento = request.form['tipo_documento']
        cedula = request.form['cedula'].strip()
        tipo_sangre = request.form['tipo_sangre'].strip().upper()
        fecha_vencimiento = request.form['fecha_vencimiento'].strip()

        iniciales = ''.join([parte[0] for parte in (nombres + ' ' + apellidos).split() if parte])
        for _ in range(10):
            codigo = f"{iniciales}{random.randint(1000, 9999)}"
            if not existe_codigo(codigo):
                break
        else:
            flash("No se pudo generar un código único. Intente nuevamente.")
            return redirect(request.url)

        datos = {
            'nombre': nombre,
            'cedula': cedula,
            'tipo_documento': tipo_documento,
            'cargo': 'Aprendiz',
            'codigo': codigo,
            'fecha_emision': hoy.strftime("%Y-%m-%d"),
            'fecha_vencimiento': fecha_vencimiento,
            'tipo_sangre': tipo_sangre,
            'foto': None
        }

        archivo_foto = request.files['foto']
        if archivo_foto and archivo_foto.filename != '':
            # Usar función con backup para aprendiz
            exito, nombre_archivo_foto, mensaje = procesar_foto_aprendiz_con_backup(archivo_foto, cedula)
            
            if exito:
                datos['foto'] = nombre_archivo_foto
            else:
                flash(f"Error procesando foto: {mensaje}")
                return redirect(request.url)
        else:
            flash("Debe subir una foto.")
            return redirect(request.url)

        try:
            insertar_empleado(datos)
            flash("Datos registrados correctamente.")
            return redirect(url_for('logout'))
        except ValueError as ve:
            flash(str(ve))
            return redirect(request.url)
        except Exception as e:
            flash(f"Error inesperado: {e}")
            return redirect(request.url)

    return render_template(
        "registro_aprendiz.html",
        usuario=session['usuario'],
        fecha_hoy=hoy.strftime("%Y-%m-%d"),
        fecha_vencimiento=vencimiento.strftime("%Y-%m-%d"),
        cedula_autenticada=obtener_cedula_aprendiz_autenticado()
    )

@app.route('/generar')
def generar():
    return generar_carnet_web()

@app.route('/generar_carnet', methods=['GET', 'POST'])
def generar_carnet_web():
    print("Ruta /generar accedida")
    print(f"Método: {request.method}")
    
    if 'usuario' not in session or session.get('rol') != 'admin':
        print("Sin autorización - redirigiendo a login")
        return redirect(url_for('login'))
    
    print(f"Usuario autorizado: {session.get('usuario')}")

    if request.method == 'POST':
        print("Procesando POST request")
        print(f"Form data completo: {dict(request.form)}")
        
        cedula = request.form.get('cedula', '').strip()
        print(f"Cédula recibida: '{cedula}'")
        
        if not cedula:
            print("Cédula vacía")
            flash("Por favor ingresa un número de cédula.", 'error')
            return render_template("generar.html")
        
        # Limpiar cédula de cualquier formato
        cedula_limpia = ''.join(filter(str.isdigit, cedula))
        print(f"Cédula limpia: '{cedula_limpia}'")
        
        if len(cedula_limpia) < 7 or len(cedula_limpia) > 10:
            print(f"Cédula inválida - longitud: {len(cedula_limpia)}")
            flash("La cédula debe tener entre 7 y 10 dígitos.", 'error')
            return render_template("generar.html")
        
        print(f"Buscando empleado con cédula: {cedula_limpia}")
        # CAMBIO PRINCIPAL: Usar la nueva función que busca con datos completos SENA
        empleado = buscar_empleado_completo(cedula_limpia)
        
        if not empleado:
            print(f"Empleado no encontrado para cédula: {cedula_limpia}")
            flash(f"No se encontró un empleado con la cédula {cedula_limpia}.", 'error')
            return render_template("generar.html")
        
        print(f"Empleado encontrado: {empleado.get('nombre', 'Sin nombre')}")
        
        try:
            print("Generando QR...")
            ruta_qr = generar_qr(empleado["cedula"])
            print(f"QR generado: {ruta_qr}")
            
            print("Generando carnet...")
            ruta_carnet = generar_carnet(empleado, ruta_qr)
            print(f"Carnet generado: {ruta_carnet}")
            
            nombre_archivo = os.path.basename(ruta_carnet)
            print(f"Nombre archivo: {nombre_archivo}")
            
            # Combinar anverso y reverso
            print("Combinando anverso y reverso...")
            reverso_path = f"reverso_{empleado['cedula']}.png"
            archivo_combinado = combinar_anverso_reverso(nombre_archivo, reverso_path, empleado['nombre'])
            print(f"Archivo combinado: {archivo_combinado}")

            # ✅ Marcar carnet como disponible SOLO si el PNG combinado existe en disco
            ruta_png = os.path.join('static', 'carnets', archivo_combinado)
            if os.path.exists(ruta_png):
                conn = sqlite3.connect('carnet.db')
                cursor = conn.cursor()
                cursor.execute(
                    "UPDATE empleados SET carnet_disponible = 1 WHERE cedula = ?",
                    (cedula_limpia,)
                )
                conn.commit()
                conn.close()
                print(f"✅ carnet_disponible = 1 para cédula {cedula_limpia}")
            else:
                print(f"⚠️ PNG combinado no encontrado en disco, no se marcó carnet_disponible")

            print("Carnet generado exitosamente!")
            flash(f"Carnet generado exitosamente para {empleado['nombre']} (Nivel: {empleado['nivel_formacion']})", 'success')
            return render_template("ver_carnet.html", carnet=archivo_combinado, empleado=empleado)
            
        except Exception as e:
            print(f"Error al generar carnet: {e}")
            print(f"Tipo de error: {type(e).__name__}")
            print(f"Traceback completo: {traceback.format_exc()}")
            flash(f"Error al generar el carné: {str(e)}", 'error')
            return render_template("generar.html")

    print("Mostrando formulario GET")
    return render_template("generar.html")

# TAREA DE VALIDACION - 16
# Descarga por número de documento: /descargar_carnet/1234567890 (Content-Disposition: attachment via as_attachment).
@app.route('/descargar_carnet/<int:cedula>')
def descargar_carnet_por_cedula(cedula):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    cedula_str = str(cedula)

    if session.get('rol') == 'aprendiz':
        cedula_autorizada = obtener_cedula_aprendiz_autenticado()
        if not cedula_autorizada:
            flash('Primero consulta tu cédula en el dashboard para vincular esta sesión.', 'error')
            return redirect(url_for('dashboard_aprendiz'))
        if cedula_str != cedula_autorizada:
            flash('No tienes permisos para descargar ese carnet.', 'error')
            return redirect(url_for('dashboard_aprendiz'))

    carnet_archivo = obtener_archivo_carnet_por_cedula(cedula_str)
    if not carnet_archivo:
        flash('El carnet no está disponible para descarga.', 'warning')
        if session.get('rol') == 'aprendiz':
            return redirect(url_for('dashboard_aprendiz'))
        return redirect(url_for('dashboard_admin'))

    return send_from_directory('static/carnets', carnet_archivo, as_attachment=True)


# TAREA DE VALIDACION - 17
# Solo el aprendiz dueño (cédula de sesión) puede descargar su archivo de carnet.
@app.route('/descargar_carnet/<path:carnet>/<cedula>')
def descargar_carnet(carnet, cedula):
    if 'usuario' not in session:
        return redirect(url_for('login'))

    if session.get('rol') == 'aprendiz':
        cedula_autorizada = obtener_cedula_aprendiz_autenticado()
        if not cedula_autorizada:
            flash('Primero consulta tu cédula en el dashboard para vincular esta sesión.', 'error')
            return redirect(url_for('dashboard_aprendiz'))

        nombre_archivo = os.path.basename(carnet)
        match_cedula = re.search(r'(\d{7,12})', nombre_archivo)
        if match_cedula:
            if match_cedula.group(1) != cedula_autorizada:
                flash('No tienes permisos para descargar ese carnet.', 'error')
                return redirect(url_for('dashboard_aprendiz'))
        else:
            # TAREA DE VALIDACION - 17
            # Permite archivos *_completo.png del aprendiz autenticado.
            carnet_autorizado = obtener_archivo_carnet_por_cedula(cedula_autorizada)
            if not carnet_autorizado or nombre_archivo != carnet_autorizado:
                flash('No tienes permisos para descargar ese carnet.', 'error')
                return redirect(url_for('dashboard_aprendiz'))
    user = buscar_empleado_completo(cedula)
    
    return send_from_directory('static/carnets', carnet, as_attachment=True, download_name = f"carnet_{user['nombre']}_{user['cedula']}.png" )


# TAREA DE VALIDACION - 17
# Descarga segura para aprendiz autenticado usando su cédula vinculada en sesión.
@app.route('/descargar_mi_carnet')
def descargar_mi_carnet():
    if 'usuario' not in session or session.get('rol') != 'aprendiz':
        flash('Debes iniciar sesión como aprendiz para descargar tu carnet.', 'error')
        return redirect(url_for('login'))

    cedula_autorizada = obtener_cedula_aprendiz_autenticado()
    if not cedula_autorizada:
        flash('Primero consulta tu cédula en el dashboard para vincular esta sesión.', 'error')
        return redirect(url_for('dashboard_aprendiz'))

    carnet_archivo = obtener_archivo_carnet_por_cedula(cedula_autorizada)
    if not carnet_archivo:
        flash('Tu carnet aún no está disponible para descarga.', 'warning')
        return redirect(url_for('dashboard_aprendiz'))

    return redirect(url_for('descargar_carnet', carnet=carnet_archivo, cedula=cedula_autorizada))

@app.route('/descargar_plantilla')
def descargar_plantilla():
    """Genera plantilla Excel con datos reales de empleados registrados"""
    if 'usuario' not in session or session['rol'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        # Obtener empleados registrados
        empleados = obtener_todos_empleados()
        
        if empleados:
            # Crear plantilla con datos reales (CON nivel_formacion y red_tecnologica)
            data = {
                'NIS': [],
                'Primer Apellido': [],
                'Segundo Apellido': [],
                'Nombre': [],
                'Tipo de documento': [],
                'Número de documento': [],
                'Tipo de Sangre': [],
                'Nombre del Programa': [],
                'Nivel de Formación': [],
                'Código de Ficha': [],
                'Centro': [],
                'Red Tecnologica': [],
                'Fecha Finalización del Programa': []
            }
            
            # Llenar con datos reales
            for empleado in empleados:
                # Dividir nombre completo en partes
                partes_nombre = empleado['nombre'].split()
                if len(partes_nombre) >= 3:
                    nombres = partes_nombre[0]
                    primer_apellido = partes_nombre[1]
                    segundo_apellido = ' '.join(partes_nombre[2:])
                elif len(partes_nombre) == 2:
                    nombres = partes_nombre[0]
                    primer_apellido = partes_nombre[1]
                    segundo_apellido = ''
                else:
                    nombres = empleado['nombre']
                    primer_apellido = ''
                    segundo_apellido = ''
                
                data['NIS'].append(empleado['nis'])
                data['Primer Apellido'].append(primer_apellido)
                data['Segundo Apellido'].append(segundo_apellido)
                data['Nombre'].append(nombres)
                data['Tipo de documento'].append(empleado['tipo_documento'])
                data['Número de documento'].append(empleado['cedula'])
                data['Tipo de Sangre'].append(empleado['tipo_sangre'])
                data['Nombre del Programa'].append(empleado['nombre_programa'])
                data['Nivel de Formación'].append(empleado['nivel_formacion'])
                data['Código de Ficha'].append(empleado['codigo_ficha'])
                data['Centro'].append(empleado['centro'])
                data['Red Tecnologica'].append(empleado['red_tecnologica'])
                data['Fecha Finalización del Programa'].append(empleado['fecha_vencimiento'])
            
            filename = f'empleados_sena_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            flash(f'Se descargó la plantilla con {len(empleados)} empleados registrados', 'success')
            
        else:
            # Plantilla con datos de ejemplo si no hay empleados
            data = {
                'NIS': ['12345678901', '12345678902', '12345678903'],
                'Primer Apellido': ['PEREZ', 'GARCIA', 'MARTINEZ'],
                'Segundo Apellido': ['LOPEZ', 'RODRIGUEZ', 'SILVA'],
                'Nombre': ['JUAN CARLOS', 'MARIA ALEJANDRA', 'CARLOS ANDRES'],
                'Tipo de documento': ['CC', 'CC', 'TI'],
                'Número de documento': ['12345678', '87654321', '11223344'],
                'Tipo de Sangre': ['O+', 'A-', 'B+'],
                'Nombre del Programa': [
                    'Análisis y Desarrollo de Sistemas de Información',
                    'Biotecnología Industrial',
                    'Gestión Empresarial'
                ],
                'Nivel de Formación': ['Técnico', 'Tecnólogo', 'Técnico'],
                'Código de Ficha': ['2024001', '2024002', '2024003'],
                'Centro': [
                    'Centro de Biotecnología Industrial',
                    'Centro de Biotecnología Industrial',
                    'Centro de Biotecnología Industrial'
                ],
                'Red Tecnologica': [
                    'Tecnologías de Producción Industrial',
                    'Tecnologías de Producción Industrial', 
                    'Gestión y Negocios'
                ],
                'Fecha Finalización del Programa': ['31/12/2024', '30/06/2025', '15/11/2024']
            }
            filename = 'plantilla_empleados_sena.xlsx'
            flash('Se descargó la plantilla con datos de ejemplo (no hay empleados registrados)', 'info')
        
        # Crear archivo Excel con openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Empleados SENA'

        headers = list(data.keys())
        header_fill = openpyxl.styles.PatternFill(start_color='39A935', end_color='39A935', fill_type='solid')
        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_align = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

        row_fill_alt = openpyxl.styles.PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
        for col_num, key in enumerate(headers, 1):
            for i, value in enumerate(data[key]):
                cell = ws.cell(row=i + 2, column=col_num, value=value)
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                if (i + 2) % 2 == 0:
                    cell.fill = row_fill_alt

        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        ws.row_dimensions[1].height = 30

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file_path = temp_file.name
        wb.save(temp_file_path)

        return send_file(temp_file_path, as_attachment=True, download_name=filename)

    except Exception as e:
        print(f"Error generando plantilla: {e}")
        flash(f'Error al generar la plantilla: {str(e)}', 'error')
        return redirect(url_for('dashboard_admin'))
        

@app.route('/cargar_plantilla', methods=['GET', 'POST'])
def cargar_plantilla():
    """Ruta MEJORADA para cargar empleados desde archivo Excel SENA"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'GET':
        return render_template('cargar_plantilla.html')
    
    if request.method == 'POST':
        try:
            # Verificar si se subió un archivo
            if 'excel_file' not in request.files:
                return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
            
            file = request.files['excel_file']
            if file.filename == '':
                return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
            
            # Verificar extensión del archivo
            if not file.filename.lower().endswith(('.xlsx', '.xls')):
                return jsonify({'success': False, 'message': 'El archivo debe ser un Excel (.xlsx o .xls)'})
            
            print(f"🔄 Procesando archivo SENA: {file.filename}")
            
            # USAR LA NUEVA FUNCIÓN MEJORADA PARA EXCEL SENA (CON VERIFICACIÓN DE DUPLICADOS)
            resultado = cargar_excel_sena_mejorado(file)
            
            print(f"✅ Resultado de carga: {resultado}")
            
            return jsonify(resultado)
                
        except Exception as e:
            error_msg = f"Error general: {str(e)}"
            print(f"❌ {error_msg}")
            return jsonify({'success': False, 'message': error_msg})

@app.route('/cargar_excel', methods=['GET', 'POST'])
def cargar_excel():
    """Alias para cargar_plantilla - compatible con el dashboard"""
    return cargar_plantilla()

@app.route('/buscar_rapido', methods=['GET', 'POST'])
def buscar_rapido():
    """Nueva ruta para búsqueda rápida de aprendices por cédula"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    aprendiz = None
    
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        
        if not cedula:
            flash('Por favor ingresa un número de cédula.', 'error')
            return render_template('buscar_rapido.html')
        
        # Limpiar cédula
        cedula_limpia = ''.join(filter(str.isdigit, cedula))
        
        if len(cedula_limpia) < 7 or len(cedula_limpia) > 10:
            flash('La cédula debe tener entre 7 y 10 dígitos.', 'error')
            return render_template('buscar_rapido.html')
        
        # Buscar empleado
        aprendiz = buscar_empleado_completo(cedula_limpia)
        
        if aprendiz:
            # Verificar si tiene foto
            if aprendiz['foto']:
                ruta_foto = os.path.join('static/fotos', aprendiz['foto'])
                aprendiz['foto_existe'] = os.path.exists(ruta_foto)
            else:
                aprendiz['foto_existe'] = False
            
            flash(f'✅ Aprendiz encontrado: {aprendiz["nombre"]}', 'success')
        else:
            flash(f'❌ No se encontró aprendiz con cédula {cedula_limpia}', 'error')
    
    return render_template('buscar_rapido.html', aprendiz=aprendiz)

# TAREA DE VALIDACION - 17
# Bloquea actualización de foto para cédulas distintas a la vinculada en sesión.
@app.route('/actualizar_foto_rapido', methods=['POST'])
def actualizar_foto_rapido():
    """Ruta para actualizar foto - accesible para aprendiz y admin"""
    if 'usuario' not in session:
        return jsonify({'success': False, 'message': 'Acceso denegado'})
    
    try:
        cedula = request.form.get('cedula', '').strip()
        archivo_foto = request.files.get('foto')
        
        if not cedula or not archivo_foto:
            return jsonify({'success': False, 'message': 'Faltan datos requeridos'})
        
        # Limpiar cédula
        cedula_limpia = ''.join(filter(str.isdigit, cedula))

        if session.get('rol') == 'aprendiz':
            cedula_autorizada = obtener_cedula_aprendiz_autenticado()
            if not cedula_autorizada:
                return jsonify({
                    'success': False,
                    'message': 'Primero consulta tu cédula en el dashboard para vincular esta sesión.'
                }), 403
            if cedula_limpia != cedula_autorizada:
                return jsonify({
                    'success': False,
                    'message': 'Solo puedes actualizar tu propia foto.'
                }), 403
        
        # Procesar foto con backup
        exito, nombre_archivo_foto, mensaje = procesar_foto_admin_con_backup(archivo_foto, cedula_limpia)
        
        if not exito:
            return jsonify({'success': False, 'message': f'Error procesando foto: {mensaje}'})
        
        # Actualizar base de datos
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        cursor.execute("UPDATE empleados SET foto = ? WHERE cedula = ?", 
                     (nombre_archivo_foto, cedula_limpia))
        
        if cursor.rowcount == 0:
            conn.close()
            return jsonify({'success': False, 'message': 'No se encontró el aprendiz para actualizar'}), 404
        
        conn.commit()
        conn.close()
        
        print(f"✅ Foto cargada para: {cedula_limpia}")
        
        return jsonify({
            'success': True, 
            'message': 'Foto cargada exitosamente. El área de administración revisará tu solicitud.',
            'foto_url': f'/static/fotos/{nombre_archivo_foto}'
        })
        
    except Exception as e:
        print(f"Error actualizando foto rápido: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/consultar_datos', methods=['GET', 'POST'])
def consultar_datos_aprendiz():
    """Ruta para que los aprendices consulten TODOS sus datos con cédula"""
    if 'usuario' not in session or session.get('rol') != 'aprendiz':
        flash('Debes iniciar sesión como aprendiz para acceder.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        
        if not cedula:
            flash('Por favor ingresa tu número de cédula.', 'error')
            return render_template('consultar_datos.html')
        
        # Limpiar cédula
        cedula_limpia = ''.join(filter(str.isdigit, cedula))
        
        # Buscar aprendiz en la base de datos con TODOS los campos
        try:
            aprendiz = buscar_empleado_completo(cedula_limpia)
            
            if aprendiz:
                # Guardar datos en sesión para el siguiente paso
                session['aprendiz_cedula'] = cedula_limpia
                session['aprendiz_datos'] = aprendiz
                
                # Mensaje de éxito
                flash(f'Datos encontrados para: {aprendiz["nombre"]}', 'success')
                
                # Renderizar template con TODOS los datos
                return render_template('consultar_datos.html', 
                                     aprendiz_encontrado=True,
                                     aprendiz=aprendiz)
            else:
                # Aprendiz no encontrado
                flash('No se encontraron tus datos en el sistema.', 'error')
                return render_template('consultar_datos.html', 
                                     no_encontrado=True,
                                     cedula_buscada=cedula_limpia)
                
        except Exception as e:
            print(f"Error consultando aprendiz: {e}")
            flash('Error al consultar los datos. Intenta de nuevo.', 'error')
            return render_template('consultar_datos.html')
    
    # GET request - mostrar formulario de búsqueda
    return render_template('consultar_datos.html')

@app.route('/cargar_foto_aprendiz', methods=['GET', 'POST'])
def cargar_foto_aprendiz():
    """Ruta para que el aprendiz cargue su foto SIN generar carnet automáticamente"""
    if 'usuario' not in session or session.get('rol') != 'aprendiz':
        flash('Debes iniciar sesión como aprendiz para acceder.', 'error')
        return redirect(url_for('login'))
    
    # Verificar que tenga datos de consulta
    aprendiz_cedula = session.get('aprendiz_cedula')
    aprendiz_datos = session.get('aprendiz_datos')
    
    if not aprendiz_cedula or not aprendiz_datos:
        flash('Primero debes consultar tus datos.', 'error')
        return redirect(url_for('consultar_datos_aprendiz'))
    
    if request.method == 'POST':
        try:
            # Validar que se subió una foto
            archivo_foto = request.files.get('foto')
            if not archivo_foto or archivo_foto.filename == '':
                flash('Debes seleccionar una foto para procesar.', 'error')
                return render_template('cargar_foto_aprendiz.html', aprendiz=aprendiz_datos)
            
            # PROCESAR LA FOTO AUTOMÁTICAMENTE CON BACKUP (3x4, fondo blanco, tamaño carnet)
            exito, nombre_archivo_foto, mensaje = procesar_foto_aprendiz_con_backup(archivo_foto, aprendiz_cedula)
            
            if not exito:
                flash(f'Error procesando la foto: {mensaje}', 'error')
                return render_template('cargar_foto_aprendiz.html', aprendiz=aprendiz_datos)
            
            print(f"Foto procesada automáticamente con backup: {nombre_archivo_foto}")
            
            # Actualizar datos del aprendiz con la nueva foto en la base de datos
            conn = sqlite3.connect('carnet.db')
            cursor = conn.cursor()
            cursor.execute("UPDATE empleados SET foto = ? WHERE cedula = ?", 
                         (nombre_archivo_foto, aprendiz_cedula))
            conn.commit()
            conn.close()
            
            # Limpiar session data
            session.pop('aprendiz_cedula', None)
            session.pop('aprendiz_datos', None)
            
            # CAMBIO PRINCIPAL: Solo mostrar mensaje de éxito, NO generar carnet
            flash('Foto subida exitosamente! Tu foto ha sido procesada y guardada con copia de respaldo. El administrador generará tu carnet pronto.', 'success')
            flash('Tu foto se procesó automáticamente con las especificaciones correctas (3x4, fondo blanco).', 'info')
            flash('Espera a que el administrador genere tu carnet. Te notificaremos cuando esté listo.', 'info')
            
            # Redirigir al dashboard del aprendiz en lugar de generar carnet
            return redirect(url_for('dashboard_aprendiz'))
            
        except Exception as e:
            print(f"Error en cargar_foto_aprendiz: {e}")
            flash(f'Error al procesar la foto: {str(e)}', 'error')
            return render_template('cargar_foto_aprendiz.html', aprendiz=aprendiz_datos)
    
    # GET request - mostrar formulario para cargar foto
    return render_template('cargar_foto_aprendiz.html', aprendiz=aprendiz_datos)

@app.route('/cancelar_consulta')
def cancelar_consulta():
    """Cancelar consulta y limpiar session"""
    session.pop('aprendiz_cedula', None)
    session.pop('aprendiz_datos', None)
    flash('Consulta cancelada.', 'info')
    return redirect(url_for('dashboard_aprendiz'))

@app.route('/consultar_aprendices')
@app.route('/admin/consultar_aprendices')
def consultar_aprendices():
    """Ruta mejorada para que el admin consulte y gestione aprendices"""
    
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Obtener parámetros de búsqueda
        buscar = request.args.get('buscar', '').strip()
        filtro_foto = request.args.get('foto', '')  # 'con_foto', 'sin_foto', '' (todos)
        filtro_programa = request.args.get('programa', '').strip()
        filtro_nivel = request.args.get('nivel', '').strip()
        filtro_ficha = request.args.get('ficha', '').strip()
        
        print(f"Parámetros de búsqueda: buscar={buscar}, foto={filtro_foto}, programa={filtro_programa}, nivel={filtro_nivel}, ficha={filtro_ficha}")
        
        # Buscar con filtros
        aprendices = buscar_empleados_con_filtros(buscar, filtro_foto, filtro_programa, filtro_nivel)
        
        # Filtro adicional por ficha si se especifica
        if filtro_ficha:
            aprendices = [a for a in aprendices if filtro_ficha in str(a.get('codigo_ficha', ''))]
        
        # Verificar existencia de fotos
        for aprendiz in aprendices:
            if aprendiz['foto']:
                ruta_foto = os.path.join('static/fotos', aprendiz['foto'])
                aprendiz['foto_existe'] = os.path.exists(ruta_foto)
            else:
                aprendiz['foto_existe'] = False
        
        # Estadísticas
        total_aprendices = len(aprendices)
        con_foto = len([a for a in aprendices if a['foto_existe']])
        sin_foto = total_aprendices - con_foto
        
        # Obtener listas para filtros
        todos_empleados = obtener_todos_empleados()
        programas = list(set([emp['nombre_programa'] for emp in todos_empleados if emp['nombre_programa']]))
        niveles = list(set([emp['nivel_formacion'] for emp in todos_empleados if emp['nivel_formacion']]))
        fichas = list(set([emp['codigo_ficha'] for emp in todos_empleados if emp['codigo_ficha']]))
        
        estadisticas = {
            'total': total_aprendices,
            'con_foto': con_foto,
            'sin_foto': sin_foto
        }
        
        filtros_data = {
            'buscar': buscar, 
            'foto': filtro_foto,
            'programa': filtro_programa,
            'nivel': filtro_nivel,
            'ficha': filtro_ficha,
            'programas': sorted(programas),
            'niveles': sorted(niveles),
            'fichas': sorted(fichas)
        }
        
        print(f"Enviando {len(aprendices)} aprendices al template")
        
        return render_template('consultar_aprendices.html', 
                             aprendices=aprendices,
                             estadisticas=estadisticas,
                             filtros=filtros_data)
        
    except Exception as e:
        print(f"Error consultando aprendices: {e}")
        flash('Error al cargar los aprendices.', 'error')
        return redirect(url_for('dashboard_admin'))

@app.route('/gestionar_fotos', methods=['GET', 'POST'])
def gestionar_fotos():
    """Ruta para gestionar fotos de aprendices"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        cedula = request.form.get('cedula', '').strip()
        
        if not cedula:
            flash('Por favor ingresa el número de cédula del aprendiz.', 'error')
            return render_template('gestionar_fotos.html')
        
        # Limpiar cédula
        cedula_limpia = ''.join(filter(str.isdigit, cedula))
        
        if len(cedula_limpia) < 7 or len(cedula_limpia) > 10:
            flash('La cédula debe tener entre 7 y 10 dígitos.', 'error')
            return render_template('gestionar_fotos.html')
        
        # Buscar aprendiz
        aprendiz = buscar_empleado_completo(cedula_limpia)
        
        if not aprendiz:
            flash(f'No se encontró aprendiz con cédula {cedula_limpia}', 'error')
            return render_template('gestionar_fotos.html')
        
        # Verificar si se está subiendo una nueva foto
        archivo_foto = request.files.get('foto')
        if archivo_foto and archivo_foto.filename != '':
            try:
                # Usar función con backup para admin
                exito, nombre_archivo_foto, mensaje = procesar_foto_admin_con_backup(archivo_foto, cedula_limpia)
                
                if exito:
                    # Actualizar base de datos
                    conn = sqlite3.connect('carnet.db')
                    cursor = conn.cursor()
                    cursor.execute("UPDATE empleados SET foto = ? WHERE cedula = ?", 
                                 (nombre_archivo_foto, cedula_limpia))
                    conn.commit()
                    conn.close()
                    
                    flash(f'Foto actualizada exitosamente para {aprendiz["nombre"]} (con backup automático)', 'success')
                    
                    # Actualizar datos del aprendiz para mostrar la nueva foto
                    aprendiz['foto'] = nombre_archivo_foto
                else:
                    flash(f'Error procesando foto: {mensaje}', 'error')
                    
            except Exception as e:
                flash(f'Error al procesar foto: {str(e)}', 'error')
        
        return render_template('gestionar_fotos.html', aprendiz=aprendiz)
    
    # GET request
    return render_template('gestionar_fotos.html')

@app.route('/admin/eliminar_foto_cedula/<cedula>', methods=['POST'])
def eliminar_foto_por_cedula(cedula):
    """Permite al admin eliminar la foto de un aprendiz por cédula"""
    
    # Verificar que el usuario sea admin
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'message': 'Acceso denegado'})
    
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Obtener la información del aprendiz por cédula
        cursor.execute("SELECT foto, nombre FROM empleados WHERE cedula = ?", (cedula,))
        resultado = cursor.fetchone()
        
        if not resultado:
            return jsonify({'success': False, 'message': f'No se encontró aprendiz con cédula {cedula}'})
        
        foto_actual, nombre_aprendiz = resultado
        
        # Eliminar archivo físico si existe
        archivos_eliminados = 0
        if foto_actual:
            # Eliminar la foto principal
            ruta_completa = os.path.join('static/fotos', foto_actual)
            if os.path.exists(ruta_completa):
                os.remove(ruta_completa)
                archivos_eliminados += 1
                print(f"Archivo eliminado: {ruta_completa}")
        
        # Buscar y eliminar otros posibles archivos de foto para esta cédula
        posibles_fotos = [
            f'foto_{cedula}.png',
            f'foto_{cedula}.jpg',
            f'foto_{cedula}.jpeg',
            f'{cedula}.png',
            f'{cedula}.jpg',
            f'{cedula}.jpeg'
        ]
        
        for posible_foto in posibles_fotos:
            ruta_posible = os.path.join('static/fotos', posible_foto)
            if os.path.exists(ruta_posible):
                os.remove(ruta_posible)
                archivos_eliminados += 1
                print(f"Archivo adicional eliminado: {ruta_posible}")
        
        # Actualizar base de datos - quitar la foto
        cursor.execute("UPDATE empleados SET foto = NULL WHERE cedula = ?", (cedula,))
        conn.commit()
        conn.close()
        
        mensaje = f'Foto eliminada exitosamente para {nombre_aprendiz}. Las copias de respaldo se mantienen intactas.'
        if archivos_eliminados > 1:
            mensaje += f' ({archivos_eliminados} archivos eliminados)'
        
        return jsonify({'success': True, 'message': mensaje})
        
    except Exception as e:
        print(f"Error al eliminar foto: {e}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/admin/eliminar_foto/<int:aprendiz_id>', methods=['POST'])
def eliminar_foto_aprendiz(aprendiz_id):
    """Permite al admin eliminar la foto de un aprendiz por ID"""
    
    # Verificar que el usuario sea admin
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo los administradores pueden eliminar fotos.', 'error')
        return redirect(url_for('dashboard_admin'))
    
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Obtener la información del aprendiz
        cursor.execute("SELECT foto, nombre, cedula FROM empleados WHERE rowid = ?", (aprendiz_id,))
        resultado = cursor.fetchone()
        
        if not resultado:
            flash('Aprendiz no encontrado.', 'error')
            return redirect(url_for('consultar_aprendices'))
        
        foto_actual, nombre_aprendiz, cedula = resultado
        
        # Eliminar archivo físico si existe
        if foto_actual:
            ruta_completa = os.path.join('static/fotos', foto_actual)
            if os.path.exists(ruta_completa):
                os.remove(ruta_completa)
                print(f"Archivo eliminado: {ruta_completa}")
        
        # Actualizar base de datos - quitar la foto
        cursor.execute("UPDATE empleados SET foto = NULL WHERE rowid = ?", (aprendiz_id,))
        conn.commit()
        conn.close()
        
        flash(f'Foto eliminada exitosamente para {nombre_aprendiz}. Las copias de respaldo se mantienen intactas. El aprendiz puede subir una nueva foto.', 'success')
        
    except Exception as e:
        print(f"Error al eliminar foto: {e}")
        flash('Error al eliminar la foto. Intenta nuevamente.', 'error')
    
    return redirect(url_for('consultar_aprendices'))

@app.route('/reporte_mensual')
def reporte_mensual():
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))

    hoy = date.today()
    try:
        mes_actual  = int(request.args.get('mes',  hoy.month))
        anio_actual = int(request.args.get('anio', hoy.year))
        if not (1 <= mes_actual <= 12):
            mes_actual = hoy.month
    except (ValueError, TypeError):
        mes_actual  = hoy.month
        anio_actual = hoy.year

    primer_dia = date(anio_actual, mes_actual, 1)
    ultimo_dia = date(anio_actual, mes_actual,
                      calendar.monthrange(anio_actual, mes_actual)[1])

    fecha_inicio_str = primer_dia.strftime('%Y-%m-%d')
    fecha_fin_str    = ultimo_dia.strftime('%Y-%m-%d')

    meses_es = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    periodo          = f"{meses_es[mes_actual-1]} {anio_actual}"
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')

    try:
        conn_anios = sqlite3.connect('carnet.db')
        cur_anios  = conn_anios.cursor()
        cur_anios.execute("SELECT MIN(fecha_emision), MAX(fecha_emision) FROM empleados")
        row_anios = cur_anios.fetchone()
        conn_anios.close()
        anio_min = int(row_anios[0][:4]) if row_anios and row_anios[0] else hoy.year
        anios_disponibles = list(range(anio_min, hoy.year + 2))
    except Exception:
        anios_disponibles = [hoy.year]

    try:
        conn   = sqlite3.connect('carnet.db')
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM empleados")
        total_aprendices = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != ''")
        con_foto_total = cursor.fetchone()[0]
        sin_foto_total = total_aprendices - con_foto_total
        pct_con_foto   = round((con_foto_total / total_aprendices * 100), 1) if total_aprendices > 0 else 0

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE fecha_emision BETWEEN ? AND ?", (fecha_inicio_str, fecha_fin_str))
        registrados_periodo = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(DISTINCT codigo_ficha) FROM empleados WHERE codigo_ficha IS NOT NULL AND codigo_ficha != ''")
        total_fichas = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(DISTINCT nombre_programa) FROM empleados WHERE nombre_programa IS NOT NULL AND nombre_programa != ''")
        total_programas = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1")
        carnets_disponibles = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 0 OR carnet_disponible IS NULL")
        carnets_no_disponibles = cursor.fetchone()[0]
        carnets_total = carnets_disponibles
        pct_carnets_disponibles    = round((carnets_disponibles    / total_aprendices * 100), 1) if total_aprendices > 0 else 0
        pct_carnets_no_disponibles = round((carnets_no_disponibles / total_aprendices * 100), 1) if total_aprendices > 0 else 0

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != '' AND (carnet_disponible = 0 OR carnet_disponible IS NULL)")
        carnets_pendientes = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND (foto IS NULL OR foto = '')")
        carnets_sin_foto = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND DATE(updated_at) BETWEEN ? AND ?", (fecha_inicio_str, fecha_fin_str))
        carnets_periodo = cursor.fetchone()[0]

        semanas = _calcular_semanas_mes(primer_dia, ultimo_dia)
        por_semana = []
        for sem in semanas:
            fi = sem['inicio'].strftime('%Y-%m-%d')
            ff = sem['fin'].strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE fecha_emision BETWEEN ? AND ?", (fi, ff))
            row_s = cursor.fetchone()
            ap = row_s[0] or 0; cf = row_s[1] or 0
            cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND DATE(updated_at) BETWEEN ? AND ?", (fi, ff))
            cs = cursor.fetchone()[0] or 0
            por_semana.append({'numero': sem['numero'], 'fecha_inicio': sem['inicio'].strftime('%d/%m'), 'fecha_fin': sem['fin'].strftime('%d/%m'), 'aprendices': ap, 'carnets': cs, 'con_foto': cf, 'sin_foto': ap - cf, 'pct_foto': round((cf / ap * 100), 1) if ap > 0 else 0})

        cursor.execute("SELECT e.codigo_ficha, COUNT(*), SUM(CASE WHEN e.foto IS NOT NULL AND e.foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN e.carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados e WHERE e.fecha_emision BETWEEN ? AND ? AND e.codigo_ficha IS NOT NULL AND e.codigo_ficha != '' GROUP BY e.codigo_ficha ORDER BY 2 DESC", (fecha_inicio_str, fecha_fin_str))
        fichas_periodo = [{'ficha': r[0], 'total': r[1], 'con_foto': r[2], 'carnets_disponibles': r[3]} for r in cursor.fetchall()]

        cursor.execute("SELECT nombre_programa, COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE nombre_programa IS NOT NULL AND nombre_programa != '' GROUP BY nombre_programa ORDER BY 2 DESC LIMIT 20")
        por_programa = [{'programa': r[0], 'total': r[1], 'con_foto': r[2], 'pct_foto': round((r[2]/r[1]*100),1) if r[1]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT COALESCE(nivel_formacion,'Sin especificar'), COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados GROUP BY nivel_formacion ORDER BY 2 DESC")
        por_nivel = [{'nivel': r[0], 'total': r[1], 'con_foto': r[2], 'carnets_disponibles': r[3], 'carnets_no_disponibles': r[1]-r[3], 'pct_total': round((r[1]/total_aprendices*100),1) if total_aprendices>0 else 0, 'pct_disponibles': round((r[3]/r[1]*100),1) if r[1]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT e.codigo_ficha, MAX(e.nombre_programa), COALESCE(MAX(e.nivel_formacion),'Técnico'), COUNT(*), SUM(CASE WHEN e.foto IS NOT NULL AND e.foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN e.carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados e WHERE e.codigo_ficha IS NOT NULL AND e.codigo_ficha != '' GROUP BY e.codigo_ficha ORDER BY 4 DESC LIMIT 15")
        top_fichas = [{'ficha': r[0], 'programa': r[1] or 'Sin programa', 'nivel': r[2], 'total': r[3], 'con_foto': r[4], 'sin_foto': r[3]-r[4], 'carnets_disponibles': r[5], 'carnets_generados': r[5], 'pct_foto': round((r[4]/r[3]*100),1) if r[3]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT codigo_ficha, MAX(nombre_programa), COUNT(*), SUM(CASE WHEN foto IS NULL OR foto = '' THEN 1 ELSE 0 END), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE codigo_ficha IS NOT NULL AND codigo_ficha != '' GROUP BY codigo_ficha HAVING SUM(CASE WHEN foto IS NULL OR foto = '' THEN 1 ELSE 0 END) > 0 ORDER BY 4 DESC")
        fichas_sin_foto = [{'ficha': r[0], 'programa': r[1] or 'Sin programa', 'total': r[2], 'sin_foto': r[3], 'con_foto': r[4], 'pct_sin_foto': round((r[3]/r[2]*100),1) if r[2]>0 else 0} for r in cursor.fetchall()]

        carnets_por_semana = []
        acumulado = 0
        for sem in semanas:
            fi = sem['inicio'].strftime('%Y-%m-%d')
            ff = sem['fin'].strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*), SUM(CASE WHEN carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados WHERE DATE(updated_at) BETWEEN ? AND ? AND carnet_disponible = 1", (fi, ff))
            row_cs = cursor.fetchone()
            gen = row_cs[0] or 0; disp = row_cs[1] or 0
            acumulado += gen
            carnets_por_semana.append({'numero': sem['numero'], 'fecha_inicio': sem['inicio'].strftime('%d/%m'), 'fecha_fin': sem['fin'].strftime('%d/%m'), 'carnets_generados': gen, 'disponibles': disp, 'no_disponibles': gen-disp, 'acumulado': acumulado, 'pct_total': round((gen/carnets_periodo*100),1) if carnets_periodo>0 else 0})

        conn.close()

        stats = {
            'total_aprendices': total_aprendices, 'registrados_periodo': registrados_periodo,
            'total_fichas': total_fichas, 'total_programas': total_programas,
            'con_foto': con_foto_total, 'sin_foto': sin_foto_total, 'pct_con_foto': pct_con_foto,
            'carnets_total': carnets_total, 'carnets_periodo': carnets_periodo,
            'carnets_disponibles': carnets_disponibles, 'carnets_no_disponibles': carnets_no_disponibles,
            'carnets_pendientes': carnets_pendientes, 'carnets_sin_foto': carnets_sin_foto,
            'pct_carnets_disponibles': pct_carnets_disponibles,
            'pct_carnets_no_disponibles': pct_carnets_no_disponibles,
            'por_semana': por_semana, 'fichas_periodo': fichas_periodo,
            'por_programa': por_programa, 'por_nivel': por_nivel,
            'top_fichas': top_fichas, 'fichas_sin_foto': fichas_sin_foto,
            'carnets_por_semana': carnets_por_semana,
        }

        return render_template(
            'reporte_mensual.html',
            stats=stats, periodo=periodo, fecha_generacion=fecha_generacion,
            fecha_inicio=primer_dia.strftime('%d/%m/%Y'),
            fecha_fin=ultimo_dia.strftime('%d/%m/%Y'),
            mes_actual=mes_actual, anio_actual=anio_actual,
            anios_disponibles=anios_disponibles,
        )

    except Exception as e:
        print(f"Error generando reporte: {e}")
        import traceback; traceback.print_exc()
        flash(f'Error al generar el reporte: {str(e)}', 'error')
        return redirect(url_for('dashboard_admin'))
import pdfkit
from flask import make_response

@app.route('/reporte_mensual/pdf')
def reporte_mensual_pdf():
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))

    hoy = date.today()
    try:
        mes_actual  = int(request.args.get('mes',  hoy.month))
        anio_actual = int(request.args.get('anio', hoy.year))
        if not (1 <= mes_actual <= 12):
            mes_actual = hoy.month
    except (ValueError, TypeError):
        mes_actual  = hoy.month
        anio_actual = hoy.year

    primer_dia = date(anio_actual, mes_actual, 1)
    ultimo_dia = date(anio_actual, mes_actual,
                      calendar.monthrange(anio_actual, mes_actual)[1])

    fecha_inicio_str = primer_dia.strftime('%Y-%m-%d')
    fecha_fin_str    = ultimo_dia.strftime('%Y-%m-%d')

    meses_es = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
    periodo          = f"{meses_es[mes_actual-1]} {anio_actual}"
    fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')

    try:
        conn_anios = sqlite3.connect('carnet.db')
        cur_anios  = conn_anios.cursor()
        cur_anios.execute("SELECT MIN(fecha_emision), MAX(fecha_emision) FROM empleados")
        row_anios = cur_anios.fetchone()
        conn_anios.close()
        anio_min = int(row_anios[0][:4]) if row_anios and row_anios[0] else hoy.year
        anios_disponibles = list(range(anio_min, hoy.year + 2))
    except Exception:
        anios_disponibles = [hoy.year]

    try:
        conn   = sqlite3.connect('carnet.db')
        cursor = conn.cursor()

        cursor.execute("SELECT COUNT(*) FROM empleados")
        total_aprendices = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != ''")
        con_foto_total = cursor.fetchone()[0]
        sin_foto_total = total_aprendices - con_foto_total
        pct_con_foto   = round((con_foto_total / total_aprendices * 100), 1) if total_aprendices > 0 else 0

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE fecha_emision BETWEEN ? AND ?", (fecha_inicio_str, fecha_fin_str))
        registrados_periodo = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(DISTINCT codigo_ficha) FROM empleados WHERE codigo_ficha IS NOT NULL AND codigo_ficha != ''")
        total_fichas = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(DISTINCT nombre_programa) FROM empleados WHERE nombre_programa IS NOT NULL AND nombre_programa != ''")
        total_programas = cursor.fetchone()[0]

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1")
        carnets_disponibles = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 0 OR carnet_disponible IS NULL")
        carnets_no_disponibles = cursor.fetchone()[0]
        carnets_total = carnets_disponibles
        pct_carnets_disponibles    = round((carnets_disponibles    / total_aprendices * 100), 1) if total_aprendices > 0 else 0
        pct_carnets_no_disponibles = round((carnets_no_disponibles / total_aprendices * 100), 1) if total_aprendices > 0 else 0

        cursor.execute("SELECT COUNT(*) FROM empleados WHERE foto IS NOT NULL AND foto != '' AND (carnet_disponible = 0 OR carnet_disponible IS NULL)")
        carnets_pendientes = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND (foto IS NULL OR foto = '')")
        carnets_sin_foto = cursor.fetchone()[0]
        cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND DATE(updated_at) BETWEEN ? AND ?", (fecha_inicio_str, fecha_fin_str))
        carnets_periodo = cursor.fetchone()[0]

        semanas = _calcular_semanas_mes(primer_dia, ultimo_dia)
        por_semana = []
        for sem in semanas:
            fi = sem['inicio'].strftime('%Y-%m-%d')
            ff = sem['fin'].strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE fecha_emision BETWEEN ? AND ?", (fi, ff))
            row_s = cursor.fetchone()
            ap = row_s[0] or 0; cf = row_s[1] or 0
            cursor.execute("SELECT COUNT(*) FROM empleados WHERE carnet_disponible = 1 AND DATE(updated_at) BETWEEN ? AND ?", (fi, ff))
            cs = cursor.fetchone()[0] or 0
            por_semana.append({'numero': sem['numero'], 'fecha_inicio': sem['inicio'].strftime('%d/%m'), 'fecha_fin': sem['fin'].strftime('%d/%m'), 'aprendices': ap, 'carnets': cs, 'con_foto': cf, 'sin_foto': ap - cf, 'pct_foto': round((cf / ap * 100), 1) if ap > 0 else 0})

        cursor.execute("SELECT e.codigo_ficha, COUNT(*), SUM(CASE WHEN e.foto IS NOT NULL AND e.foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN e.carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados e WHERE e.fecha_emision BETWEEN ? AND ? AND e.codigo_ficha IS NOT NULL AND e.codigo_ficha != '' GROUP BY e.codigo_ficha ORDER BY 2 DESC", (fecha_inicio_str, fecha_fin_str))
        fichas_periodo = [{'ficha': r[0], 'total': r[1], 'con_foto': r[2], 'carnets_disponibles': r[3]} for r in cursor.fetchall()]

        cursor.execute("SELECT nombre_programa, COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE nombre_programa IS NOT NULL AND nombre_programa != '' GROUP BY nombre_programa ORDER BY 2 DESC LIMIT 20")
        por_programa = [{'programa': r[0], 'total': r[1], 'con_foto': r[2], 'pct_foto': round((r[2]/r[1]*100),1) if r[1]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT COALESCE(nivel_formacion,'Sin especificar'), COUNT(*), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados GROUP BY nivel_formacion ORDER BY 2 DESC")
        por_nivel = [{'nivel': r[0], 'total': r[1], 'con_foto': r[2], 'carnets_disponibles': r[3], 'carnets_no_disponibles': r[1]-r[3], 'pct_total': round((r[1]/total_aprendices*100),1) if total_aprendices>0 else 0, 'pct_disponibles': round((r[3]/r[1]*100),1) if r[1]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT e.codigo_ficha, MAX(e.nombre_programa), COALESCE(MAX(e.nivel_formacion),'Técnico'), COUNT(*), SUM(CASE WHEN e.foto IS NOT NULL AND e.foto != '' THEN 1 ELSE 0 END), SUM(CASE WHEN e.carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados e WHERE e.codigo_ficha IS NOT NULL AND e.codigo_ficha != '' GROUP BY e.codigo_ficha ORDER BY 4 DESC LIMIT 15")
        top_fichas = [{'ficha': r[0], 'programa': r[1] or 'Sin programa', 'nivel': r[2], 'total': r[3], 'con_foto': r[4], 'sin_foto': r[3]-r[4], 'carnets_disponibles': r[5], 'carnets_generados': r[5], 'pct_foto': round((r[4]/r[3]*100),1) if r[3]>0 else 0} for r in cursor.fetchall()]

        cursor.execute("SELECT codigo_ficha, MAX(nombre_programa), COUNT(*), SUM(CASE WHEN foto IS NULL OR foto = '' THEN 1 ELSE 0 END), SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) FROM empleados WHERE codigo_ficha IS NOT NULL AND codigo_ficha != '' GROUP BY codigo_ficha HAVING SUM(CASE WHEN foto IS NULL OR foto = '' THEN 1 ELSE 0 END) > 0 ORDER BY 4 DESC")
        fichas_sin_foto = [{'ficha': r[0], 'programa': r[1] or 'Sin programa', 'total': r[2], 'sin_foto': r[3], 'con_foto': r[4], 'pct_sin_foto': round((r[3]/r[2]*100),1) if r[2]>0 else 0} for r in cursor.fetchall()]

        carnets_por_semana = []
        acumulado = 0
        for sem in semanas:
            fi = sem['inicio'].strftime('%Y-%m-%d')
            ff = sem['fin'].strftime('%Y-%m-%d')
            cursor.execute("SELECT COUNT(*), SUM(CASE WHEN carnet_disponible = 1 THEN 1 ELSE 0 END) FROM empleados WHERE DATE(updated_at) BETWEEN ? AND ? AND carnet_disponible = 1", (fi, ff))
            row_cs = cursor.fetchone()
            gen = row_cs[0] or 0; disp = row_cs[1] or 0
            acumulado += gen
            carnets_por_semana.append({'numero': sem['numero'], 'fecha_inicio': sem['inicio'].strftime('%d/%m'), 'fecha_fin': sem['fin'].strftime('%d/%m'), 'carnets_generados': gen, 'disponibles': disp, 'no_disponibles': gen-disp, 'acumulado': acumulado, 'pct_total': round((gen/carnets_periodo*100),1) if carnets_periodo>0 else 0})

        conn.close()

        stats = {
            'total_aprendices': total_aprendices, 'registrados_periodo': registrados_periodo,
            'total_fichas': total_fichas, 'total_programas': total_programas,
            'con_foto': con_foto_total, 'sin_foto': sin_foto_total, 'pct_con_foto': pct_con_foto,
            'carnets_total': carnets_total, 'carnets_periodo': carnets_periodo,
            'carnets_disponibles': carnets_disponibles, 'carnets_no_disponibles': carnets_no_disponibles,
            'carnets_pendientes': carnets_pendientes, 'carnets_sin_foto': carnets_sin_foto,
            'pct_carnets_disponibles': pct_carnets_disponibles,
            'pct_carnets_no_disponibles': pct_carnets_no_disponibles,
            'por_semana': por_semana, 'fichas_periodo': fichas_periodo,
            'por_programa': por_programa, 'por_nivel': por_nivel,
            'top_fichas': top_fichas, 'fichas_sin_foto': fichas_sin_foto,
            'carnets_por_semana': carnets_por_semana,
        }

        html_string = render_template(
            'reporte_mensual.html',
            stats=stats, periodo=periodo, fecha_generacion=fecha_generacion,
            fecha_inicio=primer_dia.strftime('%d/%m/%Y'),
            fecha_fin=ultimo_dia.strftime('%d/%m/%Y'),
            mes_actual=mes_actual, anio_actual=anio_actual,
            anios_disponibles=anios_disponibles,
        )

        pdf = pdfkit.from_string(html_string, False, options={
            'encoding': 'UTF-8',
            'enable-local-file-access': '',
            'quiet': '',
        })

        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=reporte_{periodo.replace(" ", "_")}.pdf'
        return response

    except Exception as e:
        print(f"Error generando PDF: {e}")
        import traceback; traceback.print_exc()
        flash(f'Error al generar el PDF: {str(e)}', 'error')
        return redirect(url_for('reporte_mensual'))
# ── Helper: divide el mes en semanas ISO ──────────────────────────────
def _calcular_semanas_mes(primer_dia, ultimo_dia):
    """
    Retorna lista de dicts {numero, inicio, fin} para cada semana
    del mes, ajustando inicio/fin al primer/último día del mes.
    """
    semanas = []
    cursor_day = primer_dia
    num_semana = 1
 
    while cursor_day <= ultimo_dia:
        # Fin de semana = domingo de esa semana o último día del mes
        fin_semana = cursor_day + timedelta(days=(6 - cursor_day.weekday()))
        if fin_semana > ultimo_dia:
            fin_semana = ultimo_dia
 
        semanas.append({
            'numero': num_semana,
            'inicio': cursor_day,
            'fin':    fin_semana,
        })
        cursor_day = fin_semana + timedelta(days=1)
        num_semana += 1
 
    return semanas
# =============================================
# RUTAS PARA GESTIÓN DE BACKUPS DE FOTOS
# =============================================

@app.route('/admin/backups_fotos')
def gestionar_backups_fotos():
    """Ruta para que el admin vea y gestione los backups de fotos"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        backups = []
        metadatos_dir = "static/fotos_backup/metadatos"
        
        if os.path.exists(metadatos_dir):
            for archivo_meta in os.listdir(metadatos_dir):
                if archivo_meta.endswith('.json'):
                    try:
                        with open(os.path.join(metadatos_dir, archivo_meta), 'r', encoding='utf-8') as f:
                            metadatos = json.load(f)
                            
                        # Verificar si el archivo de backup existe
                        metadatos['backup_existe'] = os.path.exists(metadatos['archivo_backup'])
                        
                        # Verificar si el archivo original existe
                        metadatos['original_existe'] = os.path.exists(metadatos['archivo_original'])
                        
                        # Formatear fecha para mostrar
                        fecha_obj = datetime.fromisoformat(metadatos['fecha_backup'])
                        metadatos['fecha_legible'] = fecha_obj.strftime("%d/%m/%Y %H:%M:%S")
                        
                        backups.append(metadatos)
                        
                    except Exception as e:
                        print(f"Error leyendo metadatos {archivo_meta}: {e}")
        
        # Ordenar por fecha más reciente
        backups.sort(key=lambda x: x['fecha_backup'], reverse=True)
        
        # Estadísticas
        total_backups = len(backups)
        backups_validos = len([b for b in backups if b['backup_existe']])
        por_aprendiz = len([b for b in backups if b['usuario_tipo'] == 'aprendiz'])
        por_admin = len([b for b in backups if b['usuario_tipo'] == 'admin'])
        
        estadisticas = {
            'total_backups': total_backups,
            'backups_validos': backups_validos,
            'por_aprendiz': por_aprendiz,
            'por_admin': por_admin,
            'huerfanos': total_backups - backups_validos
        }
        
        return render_template('admin_backups_fotos.html', 
                             backups=backups, 
                             stats=estadisticas)
                             
    except Exception as e:
        print(f"Error gestionando backups: {e}")
        flash('Error al cargar los backups de fotos.', 'error')
        return redirect(url_for('dashboard_admin'))

@app.route('/admin/descargar_backup_foto/<path:ruta_backup>')
def descargar_backup_foto(ruta_backup):
    """Descargar una foto de backup específica"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('login'))
    
    try:
        # Validar que la ruta esté dentro del directorio de backup
        if not ruta_backup.startswith('static/fotos_backup/'):
            flash('Ruta no válida.', 'error')
            return redirect(url_for('gestionar_backups_fotos'))
        
        if os.path.exists(ruta_backup):
            directory, filename = os.path.split(ruta_backup)
            return send_from_directory(directory, filename, as_attachment=True)
        else:
            flash('Archivo de backup no encontrado.', 'error')
            return redirect(url_for('gestionar_backups_fotos'))
            
    except Exception as e:
        print(f"Error descargando backup: {e}")
        flash('Error al descargar el backup.', 'error')
        return redirect(url_for('gestionar_backups_fotos'))

@app.route('/admin/limpiar_backups_antiguos')
def limpiar_backups_antiguos():
    """Limpiar backups de más de 6 meses"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('login'))
    
    try:
        eliminados = 0
        fecha_limite = datetime.now() - timedelta(days=180)  # 6 meses
        
        metadatos_dir = "static/fotos_backup/metadatos"
        
        if os.path.exists(metadatos_dir):
            for archivo_meta in os.listdir(metadatos_dir):
                if archivo_meta.endswith('.json'):
                    try:
                        with open(os.path.join(metadatos_dir, archivo_meta), 'r', encoding='utf-8') as f:
                            metadatos = json.load(f)
                        
                        fecha_backup = datetime.fromisoformat(metadatos['fecha_backup'])
                        
                        if fecha_backup < fecha_limite:
                            # Eliminar archivo de backup
                            if os.path.exists(metadatos['archivo_backup']):
                                os.remove(metadatos['archivo_backup'])
                            
                            # Eliminar metadatos
                            os.remove(os.path.join(metadatos_dir, archivo_meta))
                            eliminados += 1
                            
                    except Exception as e:
                        print(f"Error procesando {archivo_meta}: {e}")
        
        if eliminados > 0:
            flash(f'Se eliminaron {eliminados} backups antiguos (más de 6 meses).', 'success')
        else:
            flash('No hay backups antiguos para eliminar.', 'info')
            
    except Exception as e:
        print(f"Error limpiando backups: {e}")
        flash('Error al limpiar backups antiguos.', 'error')
    
    return redirect(url_for('gestionar_backups_fotos'))

# =============================================
# RUTAS ADICIONALES Y COMPATIBILIDAD
# =============================================

@app.route('/gestionar_aprendices')
def gestionar_aprendices():
    """Alias para consultar_aprendices"""
    return redirect(url_for('consultar_aprendices'))

@app.route('/archivo_carnets')
def archivo_carnets():
    """Ruta para mostrar carnets generados agrupados por programa o ficha"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    # Obtener parámetro de agrupación (por defecto: ficha)
    agrupar_por = request.args.get('agrupar', 'ficha')
    
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        # Obtener TODOS los aprendices con foto (listos para carnet)
        cursor.execute("""
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE cargo = 'APRENDIZ' 
              AND foto IS NOT NULL 
              AND foto != ''
            ORDER BY codigo_ficha, nombre ASC
        """)
        
        aprendices_con_foto = []
        for row in cursor.fetchall():
            aprendiz = {
                'nombre': row[0],
                'cedula': row[1],
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4],
                'fecha_emision': row[5],
                'fecha_vencimiento': row[6],
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8],
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa General',
                'codigo_ficha': row[13] or 'Sin Ficha',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico'
            }
            
            # Verificar si la foto existe físicamente
            ruta_foto = os.path.join('static/fotos', aprendiz['foto'])
            aprendiz['foto_existe'] = os.path.exists(ruta_foto)
            
            # Verificar si ya tiene carnet generado
            aprendiz['carnet_archivo'] = None
            posibles_carnets = [
                f"static/carnets/carnet_{aprendiz['cedula']}.png",
                f"static/carnets/carnet_combinado_{aprendiz['cedula']}.png",
                f"static/carnets/{aprendiz['nombre'].replace(' ', '_')}_completo.png"
            ]
            
            for carnet_path in posibles_carnets:
                if os.path.exists(carnet_path):
                    aprendiz['carnet_archivo'] = os.path.basename(carnet_path)
                    break
            
            # Solo agregar si la foto existe físicamente
            if aprendiz['foto_existe']:
                aprendices_con_foto.append(aprendiz)
        
        conn.close()
        
        print(f"📊 Total aprendices con foto: {len(aprendices_con_foto)}")
        
        # Agrupar los datos según el parámetro
        grupos = {}
        
        if agrupar_por == 'programa':
            for aprendiz in aprendices_con_foto:
                programa = aprendiz['nombre_programa']
                if programa not in grupos:
                    grupos[programa] = []
                grupos[programa].append(aprendiz)
        else:  # agrupar por ficha (DEFAULT)
            for aprendiz in aprendices_con_foto:
                ficha = aprendiz['codigo_ficha']
                if ficha not in grupos:
                    grupos[ficha] = []
                grupos[ficha].append(aprendiz)
        
        print(f"📁 Grupos creados: {len(grupos)}")
        for grupo, items in grupos.items():
            print(f"   - {grupo}: {len(items)} aprendices")
        
        # Contar carnets generados vs pendientes
        total_aprendices = len(aprendices_con_foto)
        carnets_generados = len([a for a in aprendices_con_foto if a['carnet_archivo']])
        
        # Contar por nivel de formación
        niveles_count = {}
        for aprendiz in aprendices_con_foto:
            nivel = aprendiz['nivel_formacion']
            niveles_count[nivel] = niveles_count.get(nivel, 0) + 1
        
        # Estadísticas para el template
        estadisticas = {
            'total_carnets': total_aprendices,
            'total_grupos': len(grupos),
            'carnets_generados': carnets_generados,
            'carnets_pendientes': total_aprendices - carnets_generados,
            'niveles_count': niveles_count,
            'agrupar_por': agrupar_por
        }
        
        print(f"✅ Renderizando template con {len(grupos)} grupos")
        
        return render_template('archivo_carnets.html', 
                             grupos=grupos, 
                             estadisticas=estadisticas,
                             agrupar_por=agrupar_por)
        
    except Exception as e:
        print(f"❌ ERROR en archivo_carnets: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        flash(f'Error al cargar archivo de carnets: {str(e)}', 'error')
        return redirect(url_for('dashboard_admin'))

@app.route('/ver_carnet_archivo/<cedula>')
def ver_carnet_archivo(cedula):
    """Ver un carnet específico desde el archivo"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Buscar el aprendiz por cédula
        aprendiz = buscar_empleado_completo(cedula)
        
        if not aprendiz:
            flash(f'No se encontró aprendiz con cédula {cedula}', 'error')
            return redirect(url_for('archivo_carnets'))
        
        # Buscar el archivo del carnet
        posibles_carnets = [
            f"static/carnets/{aprendiz['nombre'].replace(' ', '_')}_completo.png",
            f"static/carnets/carnet_combinado_{cedula}.png",
            f"static/carnets/carnet_{cedula}.png"
        ]
        
        carnet_encontrado = None
        for carnet_path in posibles_carnets:
            if os.path.exists(carnet_path):
                carnet_encontrado = os.path.basename(carnet_path)
                break
        
        # Si no existe el carnet, generarlo ahora
        if not carnet_encontrado:
            print(f"⚠️ Carnet no encontrado, generando ahora para {aprendiz['nombre']}")
            
            # Verificar que tenga foto
            if not aprendiz.get('foto'):
                flash(f'El aprendiz {aprendiz["nombre"]} no tiene foto. Debe cargar una primero.', 'error')
                return redirect(url_for('archivo_carnets'))
            
            try:
                # Generar QR
                ruta_qr = generar_qr(aprendiz["cedula"])
                
                # Generar carnet
                ruta_carnet = generar_carnet(aprendiz, ruta_qr)
                nombre_archivo = os.path.basename(ruta_carnet)
                
                # Combinar anverso y reverso
                reverso_path = f"reverso_{aprendiz['cedula']}.png"
                carnet_encontrado = combinar_anverso_reverso(nombre_archivo, reverso_path, aprendiz['nombre'])

                # ✅ Marcar carnet_disponible = 1 SOLO si el PNG existe en disco
                ruta_png = os.path.join('static', 'carnets', carnet_encontrado)
                if os.path.exists(ruta_png):
                    conn = sqlite3.connect('carnet.db')
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE empleados SET carnet_disponible = 1 WHERE cedula = ?",
                        (aprendiz['cedula'],)
                    )
                    conn.commit()
                    conn.close()
                    print(f"✅ carnet_disponible = 1 marcado para cédula {aprendiz['cedula']}")
                else:
                    print(f"⚠️ PNG no encontrado en disco, no se marcó carnet_disponible")
                
                flash(f'✅ Carnet generado exitosamente para {aprendiz["nombre"]}', 'success')
                
            except Exception as e:
                flash(f'Error al generar el carnet: {str(e)}', 'error')
                return redirect(url_for('archivo_carnets'))
        
        return render_template("ver_carnet.html", 
                             carnet=carnet_encontrado, 
                             empleado=aprendiz,
                             desde_archivo=True)
        
    except Exception as e:
        print(f"❌ Error viendo carnet: {e}")
        traceback.print_exc()
        flash('Error al mostrar el carnet.', 'error')
        return redirect(url_for('archivo_carnets'))

@app.route('/verificar')
def verificar():
    """Ruta para verificar carnets"""
    if 'usuario' not in session:
        return redirect(url_for('login'))
    return render_template('verificar.html')

@app.route('/verificar_carnet', methods=['POST'])
def verificar_carnet():
    """Procesar verificación de carnet por código QR"""
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    try:
        codigo_qr = request.form.get('codigo_qr', '').strip()
        
        if not codigo_qr:
            flash("Ingresa un código para verificar.", 'error')
            return redirect(url_for('verificar'))
        
        # Buscar empleado por código QR (que generalmente es la cédula)
        empleado = cargar_empleado(codigo_qr)
        
        if empleado:
            flash(f"Carnet VÁLIDO - {empleado['nombre']} (Nivel: {empleado.get('nivel_formacion', 'N/A')})", 'success')
            return render_template('verificar.html', empleado=empleado, valido=True)
        else:
            flash("Carnet NO VÁLIDO - No se encontró en el sistema", 'error')
            return render_template('verificar.html', valido=False)
            
    except Exception as e:
        print(f"Error verificando carnet: {e}")
        flash("Error al verificar el carnet.", 'error')
        return redirect(url_for('verificar'))

@app.route('/ver_carnet')
def ver_carnet():
    """Ruta para mostrar carnets generados"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return redirect(url_for('login'))
    return render_template('ver_carnet.html')

@app.route('/configuracion')
def configuracion():
    """Ruta para configuración del sistema"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    return render_template('configuracion.html')

@app.route('/reportes')
def reportes():
    """Ruta para generar reportes del sistema"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Obtener estadísticas básicas
        empleados = obtener_todos_empleados()
        total_empleados = len(empleados)
        
        # Contar por cargo
        cargos = {}
        for emp in empleados:
            cargo = emp.get('cargo', 'Sin cargo')
            cargos[cargo] = cargos.get(cargo, 0) + 1
        
        # Contar por nivel de formación
        niveles = {}
        for emp in empleados:
            nivel = emp.get('nivel_formacion', 'Sin nivel')
            niveles[nivel] = niveles.get(nivel, 0) + 1
        
        # Contar por programa
        programas = {}
        for emp in empleados:
            programa = emp.get('nombre_programa', 'Sin programa')
            programas[programa] = programas.get(programa, 0) + 1
        
        # Contar por ficha
        fichas = {}
        for emp in empleados:
            ficha = emp.get('codigo_ficha', 'Sin ficha')
            fichas[ficha] = fichas.get(ficha, 0) + 1
        
        # Empleados registrados hoy
        hoy = date.today().strftime("%Y-%m-%d")
        empleados_hoy = len([emp for emp in empleados if emp.get('fecha_emision') == hoy])
        
        # Empleados con foto
        con_foto = len([emp for emp in empleados if emp.get('foto')])
        
        estadisticas = {
            'total_empleados': total_empleados,
            'empleados_hoy': empleados_hoy,
            'con_foto': con_foto,
            'sin_foto': total_empleados - con_foto,
            'cargos': cargos,
            'niveles_formacion': niveles,
            'programas': programas,
            'fichas': fichas,
            'empleados': empleados
        }
        
        return render_template('reportes.html', stats=estadisticas)
        
    except Exception as e:
        print(f"Error generando reportes: {e}")
        flash('Error al generar reportes.', 'error')
        return redirect(url_for('dashboard_admin'))

@app.route('/dashboard_menu')
def dashboard_menu():
    """Ruta adicional para el menú del dashboard"""
    if 'usuario' not in session:
        return redirect(url_for('login'))
    
    if session.get('rol') == 'admin':
        return redirect(url_for('dashboard_admin'))
    elif session.get('rol') == 'aprendiz':
        return redirect(url_for('dashboard_aprendiz'))
    else:
        return redirect(url_for('login'))

@app.route('/gestionar_fichas')
def gestionar_fichas():
    """Ruta para gestionar aprendices por fichas"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Obtener todas las fichas con sus aprendices
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT codigo_ficha, nombre_programa, COUNT(*) as total_aprendices,
                   SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) as con_foto,
                   MIN(fecha_emision) as primera_inscripcion,
                   MAX(fecha_vencimiento) as fecha_finalizacion
            FROM empleados 
            WHERE codigo_ficha IS NOT NULL AND codigo_ficha != ''
            GROUP BY codigo_ficha, nombre_programa
            ORDER BY codigo_ficha DESC
        """)
        
        fichas = []
        for row in cursor.fetchall():
            ficha = {
                'codigo_ficha': row[0],
                'nombre_programa': row[1],
                'total_aprendices': row[2],
                'con_foto': row[3],
                'sin_foto': row[2] - row[3],
                'primera_inscripcion': row[4],
                'fecha_finalizacion': row[5],
                'porcentaje_fotos': round((row[3] / row[2]) * 100, 1) if row[2] > 0 else 0
            }
            fichas.append(ficha)
        
        conn.close()
        
        # Estadísticas generales
        total_fichas = len(fichas)
        total_aprendices = sum([f['total_aprendices'] for f in fichas])
        
        estadisticas = {
            'total_fichas': total_fichas,
            'total_aprendices': total_aprendices,
            'fichas': fichas
        }
        
        return render_template('gestionar_fichas.html', stats=estadisticas)
        
    except Exception as e:
        print(f"Error gestionando fichas: {e}")
        flash('Error al cargar las fichas.', 'error')
        return redirect(url_for('dashboard_admin'))

@app.route('/ver_ficha/<codigo_ficha>')
def ver_ficha(codigo_ficha):
    """Ver detalles de una ficha específica"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Obtener aprendices de la ficha
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE codigo_ficha = ?
            ORDER BY nombre ASC
        """, (codigo_ficha,))
        
        aprendices = []
        for row in cursor.fetchall():
            aprendiz = {
                'nombre': row[0],
                'cedula': row[1],
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4],
                'fecha_emision': row[5],
                'fecha_vencimiento': row[6],
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8],
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa General',
                'codigo_ficha': row[13] or 'Sin Ficha',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Red Tecnológica'
            }
            
            # Verificar existencia de foto
            if aprendiz['foto']:
                ruta_foto = os.path.join('static/fotos', aprendiz['foto'])
                aprendiz['foto_existe'] = os.path.exists(ruta_foto)
            else:
                aprendiz['foto_existe'] = False
            
            aprendices.append(aprendiz)
        
        conn.close()
        
        if not aprendices:
            flash(f'No se encontraron aprendices en la ficha {codigo_ficha}', 'error')
            return redirect(url_for('gestionar_fichas'))
        
        # Estadísticas de la ficha
        total_aprendices = len(aprendices)
        con_foto = len([a for a in aprendices if a['foto_existe']])
        sin_foto = total_aprendices - con_foto
        
        programa = aprendices[0]['nombre_programa'] if aprendices else 'N/A'
        centro = aprendices[0]['centro'] if aprendices else 'N/A'
        red_tecnologica = aprendices[0]['red_tecnologica'] if aprendices else 'N/A'
        
        estadisticas = {
            'codigo_ficha': codigo_ficha,
            'nombre_programa': programa,
            'centro': centro,
            'red_tecnologica': red_tecnologica,
            'total_aprendices': total_aprendices,
            'con_foto': con_foto,
            'sin_foto': sin_foto,
            'porcentaje_fotos': round((con_foto / total_aprendices) * 100, 1) if total_aprendices > 0 else 0
        }
        
        return render_template('ver_ficha.html', 
                             aprendices=aprendices, 
                             stats=estadisticas)
        
    except Exception as e:
        print(f"Error viendo ficha {codigo_ficha}: {e}")
        flash('Error al cargar la ficha.', 'error')
        return redirect(url_for('gestionar_fichas'))

@app.route('/generar_carnets_ficha/<codigo_ficha>')
def generar_carnets_ficha(codigo_ficha):
    """Generar carnets masivamente para una ficha"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        flash('Acceso denegado. Solo administradores.', 'error')
        return redirect(url_for('login'))
    
    try:
        # Obtener aprendices de la ficha que tengan foto
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT nombre, cedula, tipo_documento, cargo, codigo, 
                   fecha_emision, fecha_vencimiento, tipo_sangre, foto,
                   nis, primer_apellido, segundo_apellido, 
                   nombre_programa, codigo_ficha, centro, nivel_formacion, red_tecnologica
            FROM empleados 
            WHERE codigo_ficha = ? AND foto IS NOT NULL AND foto != ''
            ORDER BY nombre ASC
        """, (codigo_ficha,))
        
        aprendices_con_foto = []
        for row in cursor.fetchall():
            empleado = {
                'nombre': row[0],
                'cedula': row[1],
                'tipo_documento': row[2] or 'CC',
                'cargo': row[3] or 'APRENDIZ',
                'codigo': row[4],
                'fecha_emision': row[5],
                'fecha_vencimiento': row[6],
                'tipo_sangre': row[7] or 'O+',
                'foto': row[8],
                'nis': row[9] or 'N/A',
                'primer_apellido': row[10] or '',
                'segundo_apellido': row[11] or '',
                'nombre_programa': row[12] or 'Programa General',
                'codigo_ficha': row[13] or 'Sin Ficha',
                'centro': row[14] or 'Centro de Biotecnología Industrial',
                'nivel_formacion': row[15] or 'Técnico',
                'red_tecnologica': row[16] or 'Red Tecnológica'
            }
            
            # Verificar que la foto existe físicamente
            ruta_foto = os.path.join('static/fotos', empleado['foto'])
            if os.path.exists(ruta_foto):
                aprendices_con_foto.append(empleado)
        
        conn.close()
        
        if not aprendices_con_foto:
            flash(f'No hay aprendices con foto en la ficha {codigo_ficha}', 'error')
            return redirect(url_for('ver_ficha', codigo_ficha=codigo_ficha))
        
        # Generar carnets para todos los aprendices
        carnets_generados = 0
        errores = 0
        
        for empleado in aprendices_con_foto:
            try:
                # Generar QR
                ruta_qr = generar_qr(empleado["cedula"])
                
                # Generar carnet
                ruta_carnet = generar_carnet(empleado, ruta_qr)
                
                # Combinar anverso y reverso
                nombre_archivo = os.path.basename(ruta_carnet)
                reverso_path = f"reverso_{empleado['cedula']}.png"
                archivo_combinado = combinar_anverso_reverso(nombre_archivo, reverso_path, empleado['nombre'])
                
                carnets_generados += 1
                print(f"Carnet generado para: {empleado['nombre']}")
                
            except Exception as e:
                errores += 1
                print(f"Error generando carnet para {empleado['nombre']}: {e}")
        
        flash(f'Proceso completado para ficha {codigo_ficha}: {carnets_generados} carnets generados exitosamente, {errores} errores', 'success')
        
        return redirect(url_for('ver_ficha', codigo_ficha=codigo_ficha))
        
    except Exception as e:
        print(f"Error generando carnets masivos para ficha {codigo_ficha}: {e}")
        flash('Error al generar carnets masivamente.', 'error')
        return redirect(url_for('ver_ficha', codigo_ficha=codigo_ficha))

@app.route('/api/estadisticas_fichas')
def api_estadisticas_fichas():
    """API para obtener estadísticas de fichas en JSON"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT codigo_ficha, COUNT(*) as total,
                   SUM(CASE WHEN foto IS NOT NULL AND foto != '' THEN 1 ELSE 0 END) as con_foto
            FROM empleados 
            WHERE codigo_ficha IS NOT NULL AND codigo_ficha != ''
            GROUP BY codigo_ficha
            ORDER BY codigo_ficha
        """)
        
        estadisticas = []
        for row in cursor.fetchall():
            estadisticas.append({
                'ficha': row[0],
                'total': row[1],
                'con_foto': row[2],
                'sin_foto': row[1] - row[2],
                'porcentaje': round((row[2] / row[1]) * 100, 1) if row[1] > 0 else 0
            })
        
        conn.close()
        return jsonify({'success': True, 'data': estadisticas})
        
    except Exception as e:
        print(f"Error API estadísticas fichas: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/buscar_ficha/<ficha>')
def buscar_ficha(ficha):
    """Busca todos los aprendices de una ficha específica"""
    try:
        conn = sqlite3.connect('carnet.db')
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT cedula, nombre, nombre_programa as programa
            FROM empleados
            WHERE codigo_ficha = ?
            ORDER BY nombre
        """, (ficha,))
        
        aprendices = []
        for row in cursor.fetchall():
            aprendices.append({
                'cedula': row['cedula'],
                'nombre': row['nombre'],
                'programa': row['programa'] or 'Programa Técnico'
            })
        
        conn.close()
        
        return jsonify({
            'success': True,
            'aprendices': aprendices
        })
        
    except Exception as e:
        print(f"Error buscando ficha: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/api/editar_aprendiz', methods=['POST'])
def api_editar_aprendiz():
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403
    d = request.get_json()
    cedula = d.get('cedula','').strip()
    if not cedula:
        return jsonify({'success': False, 'message': 'Cédula requerida'})
    conn = sqlite3.connect('carnet.db')
    cursor = conn.cursor()
    cursor.execute("""
        UPDATE empleados SET
            nombre=?, tipo_documento=?, tipo_sangre=?, nis=?,
            codigo_ficha=?, nombre_programa=?, nivel_formacion=?,
            fecha_vencimiento=?, updated_at=CURRENT_TIMESTAMP
        WHERE cedula=?
    """, (
        d.get('nombre','').upper(), d.get('tipo_documento','CC'),
        d.get('tipo_sangre','O+'), d.get('nis',''),
        d.get('codigo_ficha',''), d.get('nombre_programa',''),
        d.get('nivel_formacion','Técnico'), d.get('fecha_vencimiento',''),
        cedula
    ))
    conn.commit(); conn.close()
    return jsonify({'success': True})


@app.route('/eliminar_ficha', methods=['POST'])
def eliminar_ficha():
    """Elimina todos los aprendices de una ficha Y sus archivos asociados"""
    try:
        data = request.get_json()
        ficha = data.get('ficha')

        if not ficha:
            return jsonify({'success': False, 'message': 'Número de ficha requerido'}), 400

        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()

        # Obtener cédulas y fotos ANTES de eliminar
        cursor.execute("""
            SELECT cedula, foto, nombre
            FROM empleados
            WHERE codigo_ficha = ?
        """, (ficha,))
        aprendices = cursor.fetchall()

        if not aprendices:
            conn.close()
            return jsonify({'success': False, 'message': 'No se encontraron aprendices en esa ficha'})

        cedulas = [row[0] for row in aprendices]

        # ── Eliminar de la BD ──
        cursor.execute("DELETE FROM empleados WHERE codigo_ficha = ?", (ficha,))
        eliminados = cursor.rowcount
        conn.commit()
        conn.close()

        # ── Limpiar archivos de fotos y carnets ──
        archivos_eliminados = 0

        for cedula, foto, nombre in aprendices:
            # Eliminar foto principal
            if foto:
                ruta_foto = os.path.join('static', 'fotos', foto)
                if os.path.exists(ruta_foto):
                    os.remove(ruta_foto)
                    archivos_eliminados += 1

            # Eliminar posibles archivos de foto por cédula
            for ext in ['png', 'jpg', 'jpeg']:
                for patron in [f'foto_{cedula}.{ext}', f'{cedula}.{ext}']:
                    ruta = os.path.join('static', 'fotos', patron)
                    if os.path.exists(ruta):
                        os.remove(ruta)
                        archivos_eliminados += 1

            # Eliminar archivos de carnet (anverso, reverso, combinado)
            carnets_a_borrar = [
                os.path.join('static', 'carnets', f'carnet_{cedula}.png'),
                os.path.join('static', 'carnets', f'reverso_{cedula}.png'),
                os.path.join('static', 'carnets', f'carnet_combinado_{cedula}.png'),
            ]
            # También el _completo.png con nombre del aprendiz
            if nombre:
                nombre_archivo = nombre.replace(' ', '_') + '_completo.png'
                carnets_a_borrar.append(os.path.join('static', 'carnets', nombre_archivo))

            for ruta_carnet in carnets_a_borrar:
                if os.path.exists(ruta_carnet):
                    os.remove(ruta_carnet)
                    archivos_eliminados += 1

            # Eliminar QR
            ruta_qr = os.path.join('static', 'qr', f'{cedula}.png')
            if os.path.exists(ruta_qr):
                os.remove(ruta_qr)
                archivos_eliminados += 1

        print(f"✅ Ficha {ficha}: {eliminados} aprendices eliminados, {archivos_eliminados} archivos borrados")

        return jsonify({
            'success': True,
            'eliminados': eliminados,
            'archivos_eliminados': archivos_eliminados
        })

    except Exception as e:
        print(f"Error eliminando ficha: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500






# =============================================
# AGREGA ESTE ENDPOINT EN TU app.py
# (puedes agregarlo junto a los otros @app.route)
# =============================================

@app.route('/api/carnets_generados')
def api_carnets_generados():
    """
    Devuelve las cédulas de aprendices que ya tienen carnet generado en disco.
    Lee los archivos en static/carnets/ y extrae las cédulas.
    """
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403

    try:
        carpeta = os.path.join('static', 'carnets')
        cedulas_con_carnet = set()

        if os.path.exists(carpeta):
            for archivo in os.listdir(carpeta):
                if not archivo.endswith('.png'):
                    continue

                # Patrón 1: carnet_1006224416.png  → cédula = 1006224416
                if archivo.startswith('carnet_') and '_completo' not in archivo:
                    cedula = archivo.replace('carnet_', '').replace('.png', '')
                    if cedula.isdigit():
                        cedulas_con_carnet.add(cedula)
                        continue

                # Patrón 2: NOMBRE_APELLIDO_completo.png
                # Buscar cédula en BD comparando con todos los empleados
                # (se resuelve con el patrón 1 para la mayoría)

                # Patrón 3: carnet_combinado_1006224416.png
                if archivo.startswith('carnet_combinado_'):
                    cedula = archivo.replace('carnet_combinado_', '').replace('.png', '')
                    if cedula.isdigit():
                        cedulas_con_carnet.add(cedula)

        # También verificar por archivos _completo.png → buscar cédula en BD
        if os.path.exists(carpeta):
            archivos_completo = [f for f in os.listdir(carpeta) if f.endswith('_completo.png')]
            if archivos_completo:
                try:
                    conn = sqlite3.connect('carnet.db')
                    cursor = conn.cursor()
                    cursor.execute("SELECT cedula, nombre FROM empleados")
                    empleados_bd = cursor.fetchall()
                    conn.close()

                    for cedula_bd, nombre_bd in empleados_bd:
                        nombre_archivo = nombre_bd.replace(' ', '_') + '_completo.png'
                        if nombre_archivo in archivos_completo:
                            cedulas_con_carnet.add(cedula_bd)
                except Exception as e:
                    print(f"Error buscando _completo.png: {e}")

        print(f"[API] Carnets generados en disco: {len(cedulas_con_carnet)}")

        return jsonify({
            'success': True,
            'total': len(cedulas_con_carnet),
            'cedulas': list(cedulas_con_carnet)
        })

    except Exception as e:
        print(f"Error en api_carnets_generados: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/eliminar_empleado/<cedula>', methods=['POST'])
def eliminar_empleado_por_cedula(cedula):
    """Elimina un aprendiz por cédula y todos sus archivos"""
    if 'usuario' not in session or session.get('rol') != 'admin':
        return jsonify({'success': False, 'message': 'Acceso denegado'}), 403

    try:
        cedula_limpia = ''.join(filter(str.isdigit, cedula))

        conn = sqlite3.connect('carnet.db')
        cursor = conn.cursor()

        cursor.execute("SELECT nombre, foto FROM empleados WHERE cedula = ?", (cedula_limpia,))
        resultado = cursor.fetchone()

        if not resultado:
            conn.close()
            return jsonify({'success': False, 'message': 'Aprendiz no encontrado'}), 404

        nombre, foto = resultado

        cursor.execute("DELETE FROM empleados WHERE cedula = ?", (cedula_limpia,))
        conn.commit()
        conn.close()

        # Limpiar todos los archivos asociados
        archivos_a_eliminar = []

        if foto:
            archivos_a_eliminar.append(os.path.join('static', 'fotos', foto))

        for ext in ['png', 'jpg', 'jpeg']:
            archivos_a_eliminar += [
                os.path.join('static', 'fotos', f'foto_{cedula_limpia}.{ext}'),
                os.path.join('static', 'fotos', f'{cedula_limpia}.{ext}'),
            ]

        archivos_a_eliminar += [
            os.path.join('static', 'carnets', f'carnet_{cedula_limpia}.png'),
            os.path.join('static', 'carnets', f'reverso_{cedula_limpia}.png'),
            os.path.join('static', 'carnets', f'carnet_combinado_{cedula_limpia}.png'),
            os.path.join('static', 'qr', f'{cedula_limpia}.png'),
        ]
        if nombre:
            archivos_a_eliminar.append(
                os.path.join('static', 'carnets', nombre.replace(' ', '_') + '_completo.png')
            )

        for archivo in archivos_a_eliminar:
            if os.path.exists(archivo):
                os.remove(archivo)

        print(f"✅ Aprendiz eliminado: {nombre} ({cedula_limpia})")
        return jsonify({'success': True, 'message': f'Aprendiz {nombre} eliminado correctamente'})

    except Exception as e:
        print(f"Error eliminando aprendiz: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500




# =============================================
# MANEJO DE ERRORES
# =============================================

@app.errorhandler(404)
def pagina_no_encontrada(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def error_interno(e):
    return render_template('500.html'), 500

@app.errorhandler(413)
def archivo_muy_grande(e):
    flash('El archivo es demasiado grande. Máximo 16MB permitido.', 'error')
    return redirect(request.url)

@app.errorhandler(Exception)
def manejar_excepcion(e):
    print(f"Error no manejado: {e}")
    print(f"Traceback: {traceback.format_exc()}")
    flash('Ha ocurrido un error inesperado. Por favor, intenta nuevamente.', 'error')
    return redirect(url_for('dashboard_admin') if session.get('rol') == 'admin' else url_for('login'))

# =============================================
# FUNCIONES DE INICIALIZACIÓN
# =============================================

def verificar_directorios():
    """Verifica y crea directorios necesarios"""
    directorios = [
        "static/fotos",
        "static/qr", 
        "static/carnets",
        "uploads",
        "templates",
        # DIRECTORIOS DE BACKUP
        "static/fotos_backup",
        "static/fotos_backup/por_fecha",
        "static/fotos_backup/metadatos"
    ]
    
    for directorio in directorios:
        if not os.path.exists(directorio):
            os.makedirs(directorio, exist_ok=True)
            print(f"Directorio creado: {directorio}")

def mostrar_estadisticas_inicio():
    """Muestra estadísticas del sistema al iniciar"""
    try:
        stats = obtener_estadisticas_dashboard()
        print("=" * 50)
        print("🏛️  SISTEMA DE CARNETIZACIÓN SENA")
        print("=" * 50)
        print(f"📊 Total de aprendices: {stats['total_aprendices']}")
        print(f"📸 Con foto: {stats['con_foto']}")
        print(f"❌ Sin foto: {stats['sin_foto']}")
        print(f"📅 Registrados hoy: {stats['registrados_hoy']}")
        print(f"📈 Esta semana: {stats['esta_semana']}")
        print("=" * 50)
        print("💾 Sistema de backup de fotos: ACTIVADO")
        print("🔗 Aplicación lista en: http://localhost:5000")
        print("=" * 50)
        
    except Exception as e:
        print(f"Error mostrando estadísticas: {e}")

# =============================================
# INICIALIZACIÓN DE LA APLICACIÓN
# =============================================

# Ejecutar funciones de inicialización
print("🚀 Iniciando Sistema de Carnetización SENA...")

# Verificar directorios (incluye los de backup)
verificar_directorios()

# Crear carpetas específicas de backup
crear_carpetas_backup()

# Actualizar base de datos
print("🔧 Verificando base de datos...")
actualizar_base_datos_sena()
print("✅ Base de datos verificada y actualizada")

# Limpiar archivos temporales
def limpiar_archivos_temporales():
    """Limpia archivos temporales antiguos"""
    try:
        temp_dir = tempfile.gettempdir()
        
        # Buscar archivos temporales de la aplicación
        for filename in os.listdir(temp_dir):
            if filename.startswith('tmp') and filename.endswith('.xlsx'):
                filepath = os.path.join(temp_dir, filename)
                try:
                    # Eliminar archivos de más de 1 hora
                    if os.path.getmtime(filepath) < time.time() - 3600:
                        os.remove(filepath)
                        print(f"Archivo temporal eliminado: {filepath}")
                except:
                    pass
                    
    except Exception as e:
        print(f"Error limpiando archivos temporales: {e}")

limpiar_archivos_temporales()

# Mostrar estadísticas
mostrar_estadisticas_inicio()

if __name__ == "__main__":
    print("🌟 Servidor Flask iniciado con sistema de backup automático")
    print("💾 Las fotos se respaldan automáticamente en static/fotos_backup/")
    app.run(debug=True, host="0.0.0.0", port=5000)